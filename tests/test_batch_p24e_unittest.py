from __future__ import annotations

import asyncio
import io
import json
from datetime import datetime, timezone
from uuid import uuid4

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from starlette.datastructures import UploadFile
from starlette.requests import Request

import litoral_trace.api.batch as batch_api
from litoral_trace.api.auth import UserTenantContext
from litoral_trace.auth.rbac import (
    Permission,
    ensure_permission,
)
from litoral_trace.services.batch import (
    BATCH_COLUMNAS,
    BATCH_SHEET_NAME,
    BATCH_XLSX_MEDIA_TYPE,
)
from litoral_trace.services.batch_imports import (
    BatchImportConflictError,
    BatchImportIdempotencyConflictError,
    BatchImportPersistenceError,
    BatchImportResult,
)
from litoral_trace.services.batch_queries import (
    BatchImportSnapshot,
)


def _request(
    path: str,
    method: str = "POST",
    *,
    content_length: int | None = None,
) -> Request:
    headers = [
        (
            b"x-request-id",
            b"p24e-request",
        ),
        (
            b"user-agent",
            b"p24e-test/1.0",
        ),
    ]

    if content_length is not None:
        headers.append(
            (
                b"content-length",
                str(content_length).encode(),
            )
        )

    return Request(
        {
            "type": "http",
            "http_version": "1.1",
            "method": method,
            "scheme": "https",
            "path": path,
            "raw_path": path.encode(),
            "query_string": b"",
            "headers": headers,
            "client": (
                "203.0.113.30",
                50000,
            ),
            "server": (
                "test",
                443,
            ),
            "root_path": "",
        }
    )


def _user(
    role: str = "admin",
    organization_id: int = 77,
) -> UserTenantContext:
    return UserTenantContext(
        user_id=10,
        username=f"{role}-user",
        organization_id=organization_id,
        organization_name=(
            f"Org {organization_id}"
        ),
        organization_slug=(
            f"org-{organization_id}"
        ),
        role=role,
        email=f"{role}@example.com",
    )


def _row(
    *,
    identificador="RODAL-001",
    proveedor="30-12345678-9",
    latitud=-27.45,
):
    return [
        identificador,
        proveedor,
        "Madera Aserrada (Pino)",
        50.0,
        latitud,
        -58.90,
        100.0,
        45.0,
    ]


def _xlsx_bytes(
    *rows,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = BATCH_SHEET_NAME
    worksheet.append(
        BATCH_COLUMNAS
    )

    for row in rows:
        worksheet.append(
            list(row)
        )

    output = io.BytesIO()
    workbook.save(
        output
    )
    workbook.close()
    return output.getvalue()


def _upload(
    payload: bytes,
    *,
    filename: str = "batch.xlsx",
) -> UploadFile:
    return UploadFile(
        filename=filename,
        file=io.BytesIO(payload),
        headers={
            "content-type": (
                BATCH_XLSX_MEDIA_TYPE
            )
        },
    )


async def _streaming_bytes(
    response,
) -> bytes:
    chunks = []

    async for chunk in response.body_iterator:
        if isinstance(
            chunk,
            str,
        ):
            chunk = chunk.encode()

        chunks.append(
            chunk
        )

    return b"".join(
        chunks
    )


def _body(
    response,
):
    return json.loads(
        response.body.decode()
    )


class FakeImportService:
    def __init__(
        self,
        result: BatchImportResult,
    ):
        self.result = result
        self.error = None
        self.kwargs = None

    def import_validated(
        self,
        validation,
        **kwargs,
    ):
        self.kwargs = {
            "validation": validation,
            **kwargs,
        }

        if self.error is not None:
            raise self.error

        return self.result


class FakeQueryService:
    def __init__(
        self,
        snapshot: BatchImportSnapshot | None,
    ):
        self.snapshot = snapshot
        self.error = None
        self.kwargs = None

    def get_by_public_id(
        self,
        **kwargs,
    ):
        self.kwargs = kwargs

        if self.error is not None:
            raise self.error

        return self.snapshot


def _result(
    *,
    replayed: bool = False,
) -> BatchImportResult:
    return BatchImportResult(
        organization_id=77,
        total_rows=1,
        inserted_rows=1,
        lote_ids=(501,),
        identifiers=("RODAL-001",),
        import_public_id=uuid4(),
        replayed=replayed,
    )


def _snapshot() -> BatchImportSnapshot:
    now = datetime.now(
        timezone.utc
    )

    return BatchImportSnapshot(
        organization_id=77,
        public_id=uuid4(),
        status="completed",
        source_filename="batch.xlsx",
        source_sha256="a" * 64,
        total_rows=1,
        inserted_rows=1,
        lote_ids=(501,),
        identifiers=("RODAL-001",),
        created_at=now,
        completed_at=now,
    )


def test_template_returns_official_xlsx():
    response = asyncio.run(
        batch_api.descargar_plantilla_excel_endpoint(
            user=_user("auditor"),
        )
    )
    payload = asyncio.run(
        _streaming_bytes(
            response
        )
    )

    workbook = load_workbook(
        io.BytesIO(payload),
        read_only=True,
    )

    try:
        worksheet = workbook[
            BATCH_SHEET_NAME
        ]
        headers = [
            cell.value
            for cell
            in next(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=1,
                )
            )
        ]

        assert headers == BATCH_COLUMNAS
        assert response.media_type == (
            BATCH_XLSX_MEDIA_TYPE
        )
        assert (
            response.headers[
                "cache-control"
            ]
            == "no-store"
        )
    finally:
        workbook.close()


def test_validate_valid_workbook_returns_preview():
    response = asyncio.run(
        batch_api.validar_batch_excel_endpoint(
            file=_upload(
                _xlsx_bytes(
                    _row()
                )
            ),
            request=_request(
                "/api/v1/batch/validate"
            ),
            user=_user(),
        )
    )
    body = _body(
        response
    )

    assert response.status_code == 200
    assert body["valid"] is True
    assert body["total_rows"] == 1
    assert body["invalid_rows"] == 0
    assert (
        body["rows"][0]["data"][
            "identificador"
        ]
        == "RODAL-001"
    )


def test_validate_semantic_errors_are_structured_200():
    response = asyncio.run(
        batch_api.validar_batch_excel_endpoint(
            file=_upload(
                _xlsx_bytes(
                    _row(
                        latitud=999.0
                    )
                )
            ),
            request=_request(
                "/api/v1/batch/validate"
            ),
            user=_user(),
        )
    )
    body = _body(
        response
    )

    assert response.status_code == 200
    assert body["valid"] is False
    assert body["invalid_rows"] == 1
    assert body["rows"][0]["errors"]
    assert (
        body["rows"][0]["errors"][0][
            "field"
        ]
        == "Latitud"
    )


def test_validate_rejects_xls():
    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.validar_batch_excel_endpoint(
                file=_upload(
                    b"not-xls",
                    filename="legacy.xls",
                ),
                request=_request(
                    "/api/v1/batch/validate"
                ),
                user=_user(),
            )
        )

    assert exc_info.value.status_code == 400
    assert (
        exc_info.value.detail["code"]
        == "UNSUPPORTED_FILE_TYPE"
    )


def test_validate_rejects_malformed_xlsx_without_raw_detail():
    secret = "SECRET_INTERNAL_EXCEPTION"

    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.validar_batch_excel_endpoint(
                file=_upload(
                    (
                        b"PK"
                        + secret.encode()
                    )
                ),
                request=_request(
                    "/api/v1/batch/validate"
                ),
                user=_user(),
            )
        )

    assert exc_info.value.status_code == 400
    rendered = json.dumps(
        exc_info.value.detail
    )
    assert secret not in rendered


def test_validate_rejects_oversized_file_before_parser(
    monkeypatch,
):
    monkeypatch.setattr(
        batch_api,
        "BATCH_MAX_FILE_BYTES",
        32,
    )

    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.validar_batch_excel_endpoint(
                file=_upload(
                    b"x" * 33
                ),
                request=_request(
                    "/api/v1/batch/validate"
                ),
                user=_user(),
            )
        )

    assert exc_info.value.status_code == 413
    assert (
        exc_info.value.detail["code"]
        == "FILE_TOO_LARGE"
    )


def test_request_content_length_over_limit_fails_closed(
    monkeypatch,
):
    parser_called = {"value": False}

    def _fail_if_parser_reached(*args, **kwargs):
        parser_called["value"] = True
        raise AssertionError(
            "parse_batch_upload_bytes must not be reached"
        )

    monkeypatch.setattr(
        batch_api,
        "parse_batch_upload_bytes",
        _fail_if_parser_reached,
    )

    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.validar_batch_excel_endpoint(
                file=_upload(
                    _xlsx_bytes(
                        _row()
                    )
                ),
                request=_request(
                    "/api/v1/batch/validate",
                    content_length=(
                        batch_api.BATCH_HTTP_MAX_REQUEST_BYTES
                        + 1
                    ),
                ),
                user=_user(),
            )
        )

    assert exc_info.value.status_code == 413
    assert (
        exc_info.value.detail["code"]
        == "REQUEST_TOO_LARGE"
    )
    assert parser_called["value"] is False
    assert "Traceback" not in str(
        exc_info.value.detail
    )


def test_import_requires_idempotency_key():
    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.importar_batch_excel_endpoint(
                file=_upload(
                    _xlsx_bytes(
                        _row()
                    )
                ),
                request=_request(
                    "/api/v1/batch/import"
                ),
                idempotency_key=None,
                user=_user(),
            )
        )

    assert exc_info.value.status_code == 400
    assert (
        exc_info.value.detail["code"]
        == "MISSING_IDEMPOTENCY_KEY"
    )


def test_import_rejects_invalid_idempotency_key():
    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.importar_batch_excel_endpoint(
                file=_upload(
                    _xlsx_bytes(
                        _row()
                    )
                ),
                request=_request(
                    "/api/v1/batch/import"
                ),
                idempotency_key="bad\nkey",
                user=_user(),
            )
        )

    assert exc_info.value.status_code == 400
    assert (
        exc_info.value.detail["code"]
        == "INVALID_IDEMPOTENCY_KEY"
    )


def test_import_semantic_failure_422_before_service(
    monkeypatch,
):
    called = {
        "service": False
    }

    def fail_if_created():
        called["service"] = True
        raise AssertionError(
            "service must not be created"
        )

    monkeypatch.setattr(
        batch_api,
        "_new_batch_import_service",
        fail_if_created,
    )

    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.importar_batch_excel_endpoint(
                file=_upload(
                    _xlsx_bytes(
                        _row(
                            latitud=999.0
                        )
                    )
                ),
                request=_request(
                    "/api/v1/batch/import"
                ),
                idempotency_key="request-1",
                user=_user(),
            )
        )

    assert exc_info.value.status_code == 422
    assert (
        exc_info.value.detail["code"]
        == "ROW_VALIDATION_FAILED"
    )
    assert called["service"] is False


def test_import_initial_returns_201_and_never_echoes_key(
    monkeypatch,
):
    service = FakeImportService(
        _result(
            replayed=False
        )
    )
    monkeypatch.setattr(
        batch_api,
        "_new_batch_import_service",
        lambda: service,
    )

    response = asyncio.run(
        batch_api.importar_batch_excel_endpoint(
            file=_upload(
                _xlsx_bytes(
                    _row()
                )
            ),
            request=_request(
                "/api/v1/batch/import"
            ),
            idempotency_key=(
                "private-idempotency-key"
            ),
            user=_user(),
        )
    )
    body = _body(
        response
    )
    rendered = json.dumps(
        body
    )

    assert response.status_code == 201
    assert body["replayed"] is False
    assert body["status"] == "completed"
    assert "private-idempotency-key" not in rendered
    assert (
        service.kwargs[
            "organization_id"
        ]
        == 77
    )
    assert (
        service.kwargs[
            "idempotency_key"
        ]
        == "private-idempotency-key"
    )


def test_import_replay_returns_200(
    monkeypatch,
):
    service = FakeImportService(
        _result(
            replayed=True
        )
    )
    monkeypatch.setattr(
        batch_api,
        "_new_batch_import_service",
        lambda: service,
    )

    response = asyncio.run(
        batch_api.importar_batch_excel_endpoint(
            file=_upload(
                _xlsx_bytes(
                    _row()
                )
            ),
            request=_request(
                "/api/v1/batch/import"
            ),
            idempotency_key="request-replay",
            user=_user(),
        )
    )

    assert response.status_code == 200
    assert _body(response)["replayed"] is True


def test_import_duplicate_conflict_maps_409(
    monkeypatch,
):
    service = FakeImportService(
        _result()
    )
    service.error = BatchImportConflictError(
        (
            "RODAL-001",
        )
    )
    monkeypatch.setattr(
        batch_api,
        "_new_batch_import_service",
        lambda: service,
    )

    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.importar_batch_excel_endpoint(
                file=_upload(
                    _xlsx_bytes(
                        _row()
                    )
                ),
                request=_request(
                    "/api/v1/batch/import"
                ),
                idempotency_key="request-dup",
                user=_user(),
            )
        )

    assert exc_info.value.status_code == 409
    assert (
        exc_info.value.detail["code"]
        == "DUPLICATE_LOTE_IDENTIFIERS"
    )


def test_import_idempotency_conflict_maps_409(
    monkeypatch,
):
    service = FakeImportService(
        _result()
    )
    service.error = (
        BatchImportIdempotencyConflictError()
    )
    monkeypatch.setattr(
        batch_api,
        "_new_batch_import_service",
        lambda: service,
    )

    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.importar_batch_excel_endpoint(
                file=_upload(
                    _xlsx_bytes(
                        _row()
                    )
                ),
                request=_request(
                    "/api/v1/batch/import"
                ),
                idempotency_key="request-conflict",
                user=_user(),
            )
        )

    assert exc_info.value.status_code == 409
    assert (
        exc_info.value.detail["code"]
        == "IDEMPOTENCY_CONFLICT"
    )


def test_import_persistence_failure_maps_safe_503(
    monkeypatch,
):
    service = FakeImportService(
        _result()
    )
    service.error = BatchImportPersistenceError(
        "SECRET DATABASE FAILURE"
    )
    monkeypatch.setattr(
        batch_api,
        "_new_batch_import_service",
        lambda: service,
    )

    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.importar_batch_excel_endpoint(
                file=_upload(
                    _xlsx_bytes(
                        _row()
                    )
                ),
                request=_request(
                    "/api/v1/batch/import"
                ),
                idempotency_key="request-db",
                user=_user(),
            )
        )

    assert exc_info.value.status_code == 503
    rendered = json.dumps(
        exc_info.value.detail
    )
    assert "SECRET DATABASE FAILURE" not in rendered


def test_status_success_and_cross_tenant_style_not_found(
    monkeypatch,
):
    snapshot = _snapshot()
    service = FakeQueryService(
        snapshot
    )
    monkeypatch.setattr(
        batch_api,
        "_new_batch_import_query_service",
        lambda: service,
    )

    response = asyncio.run(
        batch_api.obtener_batch_import_endpoint(
            public_id=snapshot.public_id,
            user=_user(
                organization_id=77
            ),
        )
    )
    body = _body(
        response
    )

    assert response.status_code == 200
    assert body["import_id"] == str(
        snapshot.public_id
    )
    assert (
        service.kwargs[
            "organization_id"
        ]
        == 77
    )

    service.snapshot = None

    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.obtener_batch_import_endpoint(
                public_id=snapshot.public_id,
                user=_user(
                    organization_id=88
                ),
            )
        )

    assert exc_info.value.status_code == 404
    assert (
        exc_info.value.detail["code"]
        == "BATCH_IMPORT_NOT_FOUND"
    )


def test_legacy_upload_is_explicitly_retired():
    with pytest.raises(
        HTTPException
    ) as exc_info:
        asyncio.run(
            batch_api.procesar_batch_excel_legacy_endpoint(
                user=_user()
            )
        )

    assert exc_info.value.status_code == 410
    assert (
        exc_info.value.detail["code"]
        == "LEGACY_BATCH_ENDPOINT_RETIRED"
    )


def test_rbac_batch_contract_is_least_privilege():
    ensure_permission(
        _user("admin"),
        Permission.LOTE_CREATE,
    )
    ensure_permission(
        _user("manager"),
        Permission.LOTE_CREATE,
    )
    ensure_permission(
        _user("auditor"),
        Permission.LOTE_READ,
    )

    for role in (
        "auditor",
        "cliente",
    ):
        with pytest.raises(
            HTTPException
        ) as exc_info:
            ensure_permission(
                _user(role),
                Permission.LOTE_CREATE,
            )

        assert exc_info.value.status_code == 403


def test_router_exposes_enterprise_contract_and_hides_legacy_upload():
    contracts = {
        (
            route.path,
            tuple(
                sorted(
                    route.methods
                    or ()
                )
            ),
            route.include_in_schema,
        )
        for route in batch_api.router.routes
    }

    assert (
        "/batch/template",
        ("GET",),
        True,
    ) in contracts
    assert (
        "/batch/validate",
        ("POST",),
        True,
    ) in contracts
    assert (
        "/batch/import",
        ("POST",),
        True,
    ) in contracts
    assert any(
        path.startswith(
            "/batch/imports/"
        )
        and methods == ("GET",)
        and in_schema is True
        for (
            path,
            methods,
            in_schema,
        ) in contracts
    )
    assert (
        "/batch/upload",
        ("POST",),
        False,
    ) in contracts
