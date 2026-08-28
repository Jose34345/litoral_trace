from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from fpdf import FPDF
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont

from litoral_trace.assurance.ingestion import (
    AssuranceIngestionValidationError,
    validate_incoming_file,
)
from litoral_trace.assurance.normalization import (
    NormalizationError,
    normalize_argentine_number,
    normalize_cuit,
    normalize_date,
    normalize_identifier,
    normalize_quantity,
)
from litoral_trace.assurance.parsers import (
    parse_csv,
    parse_pdf,
    parse_xlsx,
)
from litoral_trace.config.settings import StorageSettings


def _storage_settings() -> StorageSettings:
    return StorageSettings(max_upload_bytes=5 * 1024 * 1024)


def _xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Despachos"
    sheet.append(["Reporte exportador", None, None])
    sheet.append([None, None, None])
    sheet.append(["CUIT", "Cantidad", "Unidad"])
    sheet.append(["30-70832310-8", "1.234,50", "kg"])
    sheet.append(["TOTAL", "1.234,50", None])
    second = workbook.create_sheet("Proveedores")
    second.append(["Proveedor", "Lote"])
    second.append(["Forestal Norte", "L-001"])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _pdf_bytes(text: str | None) -> bytes:
    pdf = FPDF()
    pdf.add_page()
    if text:
        pdf.set_font("Helvetica", size=12)
        pdf.cell(0, 10, text=text)
    output = pdf.output()
    return bytes(output) if not isinstance(output, str) else output.encode("latin-1")


def _scanned_pdf_bytes() -> bytes:
    """Build a PDF with image pixels only: no searchable text layer."""
    image = Image.new("RGB", (1800, 900), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default(size=56)
    lines = (
        "FACTURA E",
        "Numero 0001 00001234",
        "Fecha 27/08/2026",
        "CUIT 30 70832310 8",
    )
    y = 90
    for line in lines:
        draw.text((90, y), line, fill="black", font=font)
        y += 170

    image_buffer = BytesIO()
    image.save(image_buffer, format="PNG")
    image.close()
    image_buffer.seek(0)

    pdf = FPDF(unit="pt", format=(1800, 900))
    pdf.add_page()
    pdf.image(image_buffer, x=0, y=0, w=1800, h=900)
    output = pdf.output()
    image_buffer.close()
    return bytes(output) if not isinstance(output, str) else output.encode("latin-1")


def test_argentine_number_normalization_supports_both_separator_orders():
    assert normalize_argentine_number("1.234,56") == Decimal("1234.56")
    assert normalize_argentine_number("1,234.56") == Decimal("1234.56")
    assert normalize_argentine_number("(2.500,00)") == Decimal("-2500.00")


def test_date_cuit_identifier_and_quantity_normalization_are_deterministic():
    assert normalize_date("27/08/2026").isoformat() == "2026-08-27"
    assert normalize_cuit("30-70832310-8") == "30708323108"
    assert normalize_identifier("  lote   ab-01 ") == "LOTE AB-01"
    quantity = normalize_quantity("1.250,50 kg")
    assert quantity.amount == Decimal("1250.50")
    assert quantity.unit == "kg"


def test_invalid_cuit_is_rejected_instead_of_silently_corrected():
    with pytest.raises(NormalizationError):
        normalize_cuit("30-70832310-3")


def test_xlsx_parser_reads_multiple_sheets_and_skips_total_rows():
    parsed = parse_xlsx(_xlsx_bytes())
    assert parsed.file_kind == "XLSX"
    assert parsed.metadata["sheet_count"] == 2
    by_name = {table.name: table for table in parsed.tables}
    assert set(by_name) == {"Despachos", "Proveedores"}
    assert len(by_name["Despachos"].rows) == 1
    assert by_name["Despachos"].rows[0]["CUIT"] == "30-70832310-8"
    assert by_name["Despachos"].source.row == 3


def test_csv_parser_detects_semicolon_and_cp1252_encoding():
    payload = "Proveedor;Cantidad;Unidad\r\nAserradero Ñandú;1.250,50;kg\r\nTOTAL;1.250,50;\r\n".encode(
        "cp1252"
    )
    parsed = parse_csv(payload)
    assert parsed.file_kind == "CSV"
    assert parsed.metadata["delimiter"] == ";"
    assert parsed.metadata["encoding"] == "cp1252"
    assert len(parsed.tables[0].rows) == 1
    assert parsed.tables[0].rows[0]["Proveedor"] == "Aserradero Ñandú"


def test_pdf_parser_extracts_digital_text_without_ocr():
    parsed = parse_pdf(_pdf_bytes("Factura E 0001-00001234"))
    assert "Factura E" in parsed.text
    assert parsed.ocr_required is False
    assert parsed.metadata["ocr_attempted"] is False
    assert parsed.metadata["ocr_applied"] is False
    assert parsed.metadata["page_count"] == 1


def test_scanned_pdf_executes_real_tesseract_ocr():
    parsed = parse_pdf(_scanned_pdf_bytes())
    assert parsed.ocr_required is False
    assert parsed.metadata["ocr_attempted"] is True
    assert parsed.metadata["ocr_applied"] is True
    assert parsed.metadata["ocr_engine"] == "tesseract"
    assert parsed.metadata["ocr_pages_processed"] == 1
    assert "FACTURA" in parsed.text.upper()


def test_blank_pdf_remains_fail_closed_when_ocr_finds_no_useful_text():
    parsed = parse_pdf(_pdf_bytes(None))
    assert parsed.ocr_required is True
    assert parsed.text == ""
    assert parsed.metadata["ocr_attempted"] is True
    assert parsed.metadata["ocr_applied"] is False


@pytest.mark.parametrize(
    ("filename", "content_type", "payload"),
    (
        ("evidencia.pdf", "application/pdf", _pdf_bytes("Documento valido 123456789")),
        (
            "operacion.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            _xlsx_bytes(),
        ),
        ("operacion.csv", "text/csv", b"Proveedor;Cantidad\nA;10\n"),
    ),
)
def test_universal_ingestion_validation_accepts_supported_real_content(
    filename: str,
    content_type: str,
    payload: bytes,
):
    validated = validate_incoming_file(
        filename=filename,
        content_type=content_type,
        content=payload,
        storage_settings=_storage_settings(),
    )
    assert validated.sha256
    assert validated.size_bytes == len(payload)


def test_universal_ingestion_validation_rejects_extension_mime_mismatch():
    with pytest.raises(AssuranceIngestionValidationError):
        validate_incoming_file(
            filename="fake.pdf",
            content_type="text/csv",
            content=b"a;b\n1;2\n",
            storage_settings=_storage_settings(),
        )


def test_universal_ingestion_validation_rejects_corrupt_xlsx():
    with pytest.raises(AssuranceIngestionValidationError):
        validate_incoming_file(
            filename="corrupto.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=b"not-a-zip",
            storage_settings=_storage_settings(),
        )
