from __future__ import annotations

from io import BytesIO
import os
import re
from uuid import uuid4

from openpyxl import load_workbook
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text

import litoral_trace.us_lacey.ingestion as ingestion_module
import litoral_trace.us_lacey.worker as worker_module
from litoral_trace.storage import (
    ObjectDeleteResult,
    ObjectHead,
    ObjectStorageNotFoundError,
    ObjectStorageStream,
    ObjectWriteResult,
)
from litoral_trace.us_lacey.db import reset_us_lacey_engine_state
from litoral_trace.us_lacey.operations import UsLaceyOperationService
from litoral_trace.us_lacey.worker import process_one_us_lacey_job
from litoral_trace.us_lacey.worker_db import reset_us_lacey_worker_engine_state
from litoral_trace.web import us_lacey_pilot_app as portal_module


pytestmark = pytest.mark.skipif(
    os.environ.get("ENABLE_POSTGRES_TESTS") != "1"
    or not os.environ.get("US_LACEY_DATABASE_URL")
    or not os.environ.get("US_LACEY_WORKER_DATABASE_URL"),
    reason="requires isolated U.S. PostgreSQL runtime and worker credentials",
)


class MemoryObjectStorage:
    """Simulated private object boundary; PostgreSQL and worker queue remain real."""

    def __init__(self) -> None:
        self.objects: dict[str, dict[str, object]] = {}
        self.put_calls = 0

    def put_object(self, *, key, body, content_type, content_length, metadata=None):
        payload = body if isinstance(body, bytes) else body.read()
        assert len(payload) == content_length
        self.put_calls += 1
        self.objects[key] = {
            "body": payload,
            "content_type": content_type,
            "metadata": dict(metadata or {}),
            "etag": f"etag-{self.put_calls}",
            "version_id": None,
        }
        return ObjectWriteResult(etag=f"etag-{self.put_calls}", version_id=None)

    def _head(self, key: str, version_id=None) -> ObjectHead:
        item = self.objects.get(key)
        if item is None:
            raise ObjectStorageNotFoundError("head_object")
        payload = item["body"]
        return ObjectHead(
            size_bytes=len(payload),
            content_type=str(item["content_type"]),
            etag=str(item["etag"]),
            version_id=version_id,
            metadata=dict(item["metadata"]),
        )

    def head_object(self, *, key, version_id=None):
        return self._head(key, version_id)

    def get_object_stream(self, *, key, version_id=None):
        item = self.objects.get(key)
        if item is None:
            raise ObjectStorageNotFoundError("get_object")
        return ObjectStorageStream(
            body=BytesIO(item["body"]),
            head=self._head(key, version_id),
        )

    def delete_object(self, *, key, version_id=None):
        self.objects.pop(key, None)
        return ObjectDeleteResult(delete_marker=False, version_id=version_id)

    def object_exists(self, *, key, version_id=None):
        return key in self.objects

    def health_check(self):
        return True


def _configure(monkeypatch: pytest.MonkeyPatch) -> None:
    values = {
        "US_LACEY_ENVIRONMENT": "test",
        "US_LACEY_STORAGE_BUCKET": "us-lacey-ci-private",
        "US_LACEY_STORAGE_PREFIX": "us-lacey/ci-active-http-e2e",
        "US_LACEY_APP_HOSTNAME": "app.lacey.litoraltrace.com",
        "US_LACEY_SESSION_TTL_HOURS": "1",
        "US_LACEY_PRIVATE_BETA_PRICE_CENTS": "12500",
        # Deliberately one slot: upload/review/export must keep working after it is consumed.
        "US_LACEY_MONTHLY_OPERATION_LIMIT": "1",
        "US_LACEY_PAYMENT_PROVIDER": "WISE",
        "US_LACEY_BANK_TRANSFER_INSTRUCTIONS": "CI-only Wise USD transfer instructions",
        "US_LACEY_TERMS_VERSION": "terms-active-e2e-v1",
        "US_LACEY_PRIVACY_VERSION": "privacy-active-e2e-v1",
        "US_LACEY_BETA_TERMS_VERSION": "beta-active-e2e-v1",
        "US_LACEY_SUPPORT_EMAIL": "support@litoraltrace.com",
        "US_LACEY_TERMS_URL": "https://lacey.litoraltrace.com/terms",
        "US_LACEY_PRIVACY_URL": "https://lacey.litoraltrace.com/privacy",
        "US_LACEY_BETA_TERMS_URL": "https://lacey.litoraltrace.com/private-beta-terms",
    }
    for name, value in values.items():
        monkeypatch.setenv(name, value)
    monkeypatch.delenv("DATABASE_URL", raising=False)


def _organization_id_for_email(email: str) -> int:
    runtime = create_engine(
        os.environ["US_LACEY_DATABASE_URL"],
        pool_pre_ping=True,
        hide_parameters=True,
    )
    try:
        with runtime.connect() as connection:
            row = connection.execute(
                text("SELECT * FROM public.us_lacey_portal_login_lookup(:email)"),
                {"email": email},
            ).mappings().one()
            return int(row["organization_id"])
    finally:
        runtime.dispose()


def _activate_account(organization_id: int) -> None:
    """Test-only commercial activation through the tenant-scoped runtime RLS role."""
    runtime = create_engine(
        os.environ["US_LACEY_DATABASE_URL"],
        pool_pre_ping=True,
        hide_parameters=True,
    )
    try:
        with runtime.begin() as connection:
            connection.execute(
                text("SELECT set_config('app.current_organization_id', :organization_id, true)"),
                {"organization_id": str(organization_id)},
            )
            assert connection.execute(
                text(
                    "UPDATE public.us_lacey_organization_profiles "
                    "SET account_status='ACTIVE', updated_at=now() "
                    "WHERE organization_id=:organization_id RETURNING id"
                ),
                {"organization_id": organization_id},
            ).scalar_one_or_none() is not None
            assert connection.execute(
                text(
                    "UPDATE public.us_lacey_subscriptions "
                    "SET status='ACTIVE', started_at=coalesce(started_at, now()), updated_at=now() "
                    "WHERE organization_id=:organization_id RETURNING id"
                ),
                {"organization_id": organization_id},
            ).scalar_one_or_none() is not None
            assert connection.execute(
                text(
                    "UPDATE public.us_lacey_payments "
                    "SET status='VERIFIED', verified_at=now(), paid_at=now(), updated_at=now() "
                    "WHERE organization_id=:organization_id RETURNING id"
                ),
                {"organization_id": organization_id},
            ).scalar_one_or_none() is not None
    finally:
        runtime.dispose()


def _csrf_for(html: str, action: str) -> str:
    pattern = re.compile(
        rf'<form[^>]*action="{re.escape(action)}"[^>]*>.*?name="csrf_token" value="([a-f0-9]+)"',
        re.DOTALL,
    )
    match = pattern.search(html)
    assert match is not None, f"CSRF token not rendered for {action}"
    return match.group(1)


def _assert_href(html: str, href: str) -> None:
    """Assert stable navigation contracts without coupling E2E to display copy."""
    assert f'href="{href}"' in html, f"Link target not rendered: {href}"


def _csv_bytes() -> bytes:
    return (
        "HTS Code,Merchandise Description,Genus,Species,Country of Harvest,Plant Quantity,Metric Unit,Bill of Lading\n"
        "4407.11,Pine boards,Pinus,Pinus taeda,Brazil,1000,KG,BOL-E2E-9001\n"
    ).encode("utf-8")


_REQUIRED_REVIEW_VALUES = {
    "estimated_arrival_date": "2026-09-15",
    "filing_entry_reference": "123-4567890-1",
    "manufacturer_id": "USWOOD12345",
    "importer_address": "100 Test Ave, Miami FL 33101",
    "consignee_address": "200 Test Blvd, Savannah GA 31401",
    "entered_value": "12500",
    "article_component": "Pine boards",
    "percent_recycled": "0",
}


def test_active_customer_operations_upload_review_complete_exports_and_history(
    monkeypatch: pytest.MonkeyPatch,
):
    """Full ACTIVE browser journey with real PostgreSQL, queue and worker processing.

    This integration test protects behavioral contracts (HTTP transitions,
    persisted operation state, queue/worker execution and export artifacts).
    Human-facing wording is covered separately by UI/template/visual tests and
    is intentionally not used here as a proxy for application correctness.
    """
    _configure(monkeypatch)
    reset_us_lacey_engine_state()
    reset_us_lacey_worker_engine_state()
    storage = MemoryObjectStorage()
    monkeypatch.setattr(ingestion_module, "get_us_lacey_storage_client", lambda: storage)
    monkeypatch.setattr(worker_module, "get_us_lacey_storage_client", lambda: storage)

    delivered: dict[str, str] = {}

    def capture_verification_email(**kwargs) -> None:
        delivered.update({key: str(value) for key, value in kwargs.items()})

    monkeypatch.setattr(
        portal_module,
        "send_us_lacey_verification_email",
        capture_verification_email,
    )

    suffix = uuid4().hex[:12]
    email = f"active-http-e2e-{suffix}@example.com"
    password = "correct-horse-active-http-e2e-123"
    legal_name = f"Active HTTP E2E Imports {suffix} LLC"
    reference = f"ACTIVE-E2E-{suffix}"

    with TestClient(portal_module.app, follow_redirects=False) as client:
        signup = client.post(
            "/signup",
            data={
                "legal_name": legal_name,
                "business_type": "IMPORTER",
                "admin_name": "Active HTTP E2E Admin",
                "admin_email": email,
                "password": password,
                "accept_terms": "yes",
                "accept_privacy": "yes",
                "accept_beta": "yes",
            },
        )
        assert signup.status_code == 201
        token = delivered["verification_token"]

        verified = client.get(f"/verify-email?token={token}")
        assert verified.status_code == 303
        organization_id = _organization_id_for_email(email)
        _activate_account(organization_id)

        login = client.post("/login", data={"email": email, "password": password})
        assert login.status_code == 303
        assert login.headers["location"] == "/operations"
        assert client.cookies.get("us_lacey_session")

        operations = client.get("/operations")
        assert operations.status_code == 200
        assert legal_name in operations.text
        _assert_href(operations.text, "/operations/new")

        new_page = client.get("/operations/new")
        assert new_page.status_code == 200
        create_csrf = _csrf_for(new_page.text, "/operations/new")
        created = client.post(
            "/operations/new",
            data={
                "csrf_token": create_csrf,
                "client_reference": reference,
                "importer_name": legal_name,
                "supplier_name": "Brazil Pine Supplier SA",
                "consignee_name": "E2E Consignee",
                "broker_name": "E2E Customs Broker",
                "operation_date": "2026-08-30",
                "line_references": "1",
            },
        )
        assert created.status_code == 303
        operation_path = created.headers["location"]
        assert operation_path.startswith("/operations/")
        operation_public_id = operation_path.rsplit("/", 1)[-1]

        operation_service = UsLaceyOperationService()
        initial_detail = operation_service.get_detail(
            organization_id=organization_id,
            operation_public_id=operation_public_id,
        )
        assert initial_detail.document_count == 0
        assert initial_detail.status != "COMPLETED"

        # The only slot is now consumed. Creating another operation is blocked,
        # but the existing operation must remain fully usable.
        after_create = client.get("/operations")
        assert after_create.status_code == 200
        assert reference in after_create.text
        blocked_new = client.get("/operations/new")
        assert blocked_new.status_code == 409

        detail = client.get(operation_path)
        assert detail.status_code == 200
        upload_action = f"{operation_path}/upload"
        upload_csrf = _csrf_for(detail.text, upload_action)

        uploaded = client.post(
            upload_action,
            data={
                "csrf_token": upload_csrf,
                "document_role": "SUPPLIER_SHEET",
            },
            files={"document": ("shipment.csv", _csv_bytes(), "text/csv")},
        )
        assert uploaded.status_code == 303
        assert uploaded.headers["location"] == f"{operation_path}?uploaded=1"
        assert storage.put_calls == 1

        worker_result = process_one_us_lacey_job(worker_id=f"active-e2e-{suffix}")
        assert worker_result.claimed is True
        assert worker_result.job_status == "COMPLETED"
        assert worker_result.document_status in {"EXTRACTED", "NEEDS_REVIEW"}

        review_page = client.get(operation_path)
        assert review_page.status_code == 200
        assert "shipment.csv" in review_page.text
        assert "Country of Harvest" in review_page.text
        assert "Brazil" in review_page.text

        operation_detail = operation_service.get_detail(
            organization_id=organization_id,
            operation_public_id=operation_public_id,
        )
        assert operation_detail.document_count == 1
        exceptions = [
            field
            for field in operation_detail.fields
            if field.status in {"MISSING", "REVIEW"}
        ]
        assert exceptions

        # Every human decision is submitted through the actual browser route and
        # the CSRF token rendered for that exact operation+field form. Missing
        # PPQ fields receive explicit values that satisfy their field contract;
        # the test must never bypass production validation with generic strings.
        for field in exceptions:
            review_action = f"{operation_path}/review/{field.id}"
            field_csrf = _csrf_for(review_page.text, review_action)
            if field.proposed_value:
                payload = {
                    "csrf_token": field_csrf,
                    "action": "accept",
                    "value": "",
                }
            else:
                assert field.field_name in _REQUIRED_REVIEW_VALUES, (
                    f"Unexpected missing PPQ field without an explicit E2E value: {field.field_name}"
                )
                payload = {
                    "csrf_token": field_csrf,
                    "action": "edit",
                    "value": _REQUIRED_REVIEW_VALUES[field.field_name],
                }
            reviewed = client.post(review_action, data=payload)
            assert reviewed.status_code == 303
            assert reviewed.headers["location"] == operation_path

        ready_to_complete = client.get(operation_path)
        assert ready_to_complete.status_code == 200
        ready_detail = operation_service.get_detail(
            organization_id=organization_id,
            operation_public_id=operation_public_id,
        )
        unresolved = [
            field
            for field in ready_detail.fields
            if field.status in {"MISSING", "REVIEW"}
        ]
        assert unresolved == []

        complete_action = f"{operation_path}/complete"
        complete_csrf = _csrf_for(ready_to_complete.text, complete_action)
        completed = client.post(complete_action, data={"csrf_token": complete_csrf})
        assert completed.status_code == 303
        assert completed.headers["location"] == f"{operation_path}?completed=1"

        completed_detail = operation_service.get_detail(
            organization_id=organization_id,
            operation_public_id=operation_public_id,
        )
        assert completed_detail.status == "COMPLETED"

        complete_page = client.get(operation_path)
        assert complete_page.status_code == 200
        _assert_href(complete_page.text, f"{operation_path}/export.xlsx")
        _assert_href(complete_page.text, f"{operation_path}/export.csv")
        # This safety disclaimer is itself a product contract and should remain visible.
        assert "not a legal compliance determination" in complete_page.text
        assert "ACE/LAWGS" in complete_page.text

        csv_export = client.get(f"{operation_path}/export.csv")
        assert csv_export.status_code == 200
        assert "text/csv" in csv_export.headers["content-type"]
        assert "attachment;" in csv_export.headers["content-disposition"]
        assert b"Line Reference" in csv_export.content
        assert b"Brazil" in csv_export.content

        xlsx_export = client.get(f"{operation_path}/export.xlsx")
        assert xlsx_export.status_code == 200
        assert "spreadsheetml.sheet" in xlsx_export.headers["content-type"]
        workbook = load_workbook(BytesIO(xlsx_export.content), read_only=True, data_only=True)
        try:
            assert "Read Me" in workbook.sheetnames
            assert "Preparation Data" in workbook.sheetnames
            assert "Evidence" in workbook.sheetnames
            assert "Exceptions" in workbook.sheetnames
        finally:
            workbook.close()

        logout = client.post("/logout")
        assert logout.status_code == 303
        assert logout.headers["location"] == "/login"
        assert client.get("/operations").headers["location"] == "/login"

        relogin = client.post("/login", data={"email": email, "password": password})
        assert relogin.status_code == 303
        assert relogin.headers["location"] == "/operations"
        history = client.get("/operations")
        assert history.status_code == 200
        assert reference in history.text
        historical_detail = client.get(operation_path)
        assert historical_detail.status_code == 200
        _assert_href(historical_detail.text, f"{operation_path}/export.xlsx")
        _assert_href(historical_detail.text, f"{operation_path}/export.csv")
