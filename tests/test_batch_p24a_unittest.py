from __future__ import annotations

import hashlib
import io
import re
import zipfile

from openpyxl import Workbook, load_workbook
import pytest

import litoral_trace.services.batch as batch_service
from litoral_trace.services.batch import (
    BATCH_COLUMNAS,
    BATCH_MAX_FILE_BYTES,
    BATCH_MAX_ROWS,
    BATCH_SHEET_NAME,
    BatchExcelValidationError,
    generar_plantilla_excel,
    normalizar_nombre_archivo_batch,
    parsear_excel_lotes,
)


def _valid_row(index: int = 1) -> list[object]:
    return [
        f"RODAL-{index:03d}",
        f"30-1234567{index % 10}-9",
        "Madera Aserrada (Pino)",
        50.0,
        -27.45,
        -58.90,
        100.0,
        45.0,
    ]


def _workbook_bytes(
    *,
    headers: list[object] | None = None,
    rows: list[list[object]] | None = None,
    sheet_name: str = BATCH_SHEET_NAME,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(list(headers or BATCH_COLUMNAS))

    actual_rows = [_valid_row()] if rows is None else rows
    for row in actual_rows:
        worksheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _with_extra_zip_member(
    payload: bytes,
    *,
    name: str,
    body: bytes,
) -> bytes:
    source = zipfile.ZipFile(io.BytesIO(payload), "r")
    buffer = io.BytesIO()

    with source, zipfile.ZipFile(
        buffer,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as target:
        for info in source.infolist():
            target.writestr(
                info,
                source.read(info.filename),
            )
        target.writestr(name, body)

    return buffer.getvalue()


def _remove_sheet_dimension(
    payload: bytes,
) -> bytes:
    source = zipfile.ZipFile(io.BytesIO(payload), "r")
    buffer = io.BytesIO()
    dimension_pattern = re.compile(
        rb"<dimension[^>]*/>"
    )

    with source, zipfile.ZipFile(
        buffer,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as target:
        for info in source.infolist():
            body = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                body = dimension_pattern.sub(
                    b"",
                    body,
                    count=1,
                )
            target.writestr(
                info,
                body,
            )

    return buffer.getvalue()


def _rewrite_member(
    payload: bytes,
    *,
    name: str,
    transform,
) -> bytes:
    source = zipfile.ZipFile(io.BytesIO(payload), "r")
    buffer = io.BytesIO()

    with source, zipfile.ZipFile(
        buffer,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as target:
        for info in source.infolist():
            body = source.read(info.filename)
            if info.filename == name:
                body = transform(body)
            target.writestr(
                info,
                body,
            )

    return buffer.getvalue()


def _assert_code(
    exc_info: pytest.ExceptionInfo[BatchExcelValidationError],
    expected: str,
) -> None:
    assert exc_info.value.code == expected
    assert exc_info.value.detail
    assert "Traceback" not in exc_info.value.detail


def test_official_template_round_trips_through_secure_parser():
    payload = generar_plantilla_excel()

    parsed = parsear_excel_lotes(
        payload,
        filename="LitoralTrace_Plantilla_Ingreso.xlsx",
    )

    assert parsed.filename == "LitoralTrace_Plantilla_Ingreso.xlsx"
    assert parsed.sheet_name == BATCH_SHEET_NAME
    assert parsed.row_count == 1
    assert list(parsed.dataframe.columns) == BATCH_COLUMNAS
    assert parsed.sha256 == hashlib.sha256(payload).hexdigest()
    assert (
        parsed.dataframe.iloc[0]["Identificador_Lote"]
        == "Rodal_Norte_01"
    )


def test_filename_is_reduced_to_safe_basename_and_xlsx_only():
    assert (
        normalizar_nombre_archivo_batch(
            r"C:\Users\operator\Importacion.xlsx"
        )
        == "Importacion.xlsx"
    )

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        normalizar_nombre_archivo_batch("legacy.xls")

    _assert_code(exc_info, "UNSUPPORTED_FILE_TYPE")


def test_filename_nfkc_and_control_characters_fail_closed_or_normalize_safely():
    assert (
        normalizar_nombre_archivo_batch(
            "Ｆｏｌｄｅｒ／Importación_Árbol.xlsx"
        )
        == "Importación_Árbol.xlsx"
    )

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        normalizar_nombre_archivo_batch(
            "bad\nname.xlsx"
        )

    _assert_code(exc_info, "INVALID_FILENAME")


def test_empty_and_oversized_files_fail_closed():
    with pytest.raises(
        BatchExcelValidationError
    ) as empty_exc:
        parsear_excel_lotes(
            b"",
            filename="batch.xlsx",
        )

    _assert_code(empty_exc, "EMPTY_FILE")

    with pytest.raises(
        BatchExcelValidationError
    ) as size_exc:
        parsear_excel_lotes(
            b"P" * (BATCH_MAX_FILE_BYTES + 1),
            filename="batch.xlsx",
        )

    _assert_code(size_exc, "FILE_TOO_LARGE")


def test_non_zip_payload_is_rejected_before_openpyxl():
    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            b"not-an-xlsx",
            filename="batch.xlsx",
        )

    _assert_code(exc_info, "INVALID_XLSX_CONTAINER")


def test_zip_without_required_xlsx_structure_is_rejected():
    buffer = io.BytesIO()
    with zipfile.ZipFile(
        buffer,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as archive:
        archive.writestr(
            "[Content_Types].xml",
            b"<Types></Types>",
        )

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            buffer.getvalue(),
            filename="batch.xlsx",
        )

    _assert_code(exc_info, "INVALID_XLSX_STRUCTURE")


def test_required_sheet_name_is_enforced():
    payload = _workbook_bytes(sheet_name="Datos")

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            payload,
            filename="batch.xlsx",
        )

    _assert_code(exc_info, "MISSING_REQUIRED_SHEET")


@pytest.mark.parametrize(
    "headers",
    [
        BATCH_COLUMNAS[:-1],
        BATCH_COLUMNAS[:-1] + ["Columna_Inesperada"],
        [
            "Identificador_Lote",
            "ID_Proveedor",
            "Producto_Forestal",
            "Hectareas",
            "Latitud",
            "Longitud",
            "Volumen_Ingresado_Ton",
            "Volumen_Ingresado_Ton",
        ],
    ],
)
def test_schema_drift_and_duplicate_headers_are_rejected(
    headers: list[str],
):
    payload = _workbook_bytes(headers=headers)

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            payload,
            filename="batch.xlsx",
        )

    assert exc_info.value.code in {
        "DUPLICATE_HEADERS",
        "INVALID_HEADERS",
    }


def test_formula_cells_are_rejected():
    row = _valid_row()
    row[3] = "=25+25"
    payload = _workbook_bytes(rows=[row])

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            payload,
            filename="batch.xlsx",
        )

    _assert_code(exc_info, "FORMULA_NOT_ALLOWED")


def test_external_or_embedded_xlsx_features_are_rejected():
    payload = _with_extra_zip_member(
        _workbook_bytes(),
        name="xl/externalLinks/externalLink1.xml",
        body=b"<externalLink/>",
    )

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            payload,
            filename="batch.xlsx",
        )

    _assert_code(exc_info, "UNSUPPORTED_XLSX_FEATURE")


def test_doctype_and_entity_declarations_are_rejected():
    payload = _with_extra_zip_member(
        _workbook_bytes(),
        name="customXml/item1.xml",
        body=(
            b'<?xml version="1.0"?>'
            b'<!DOCTYPE x [<!ENTITY e "x">]>'
            b"<x>&e;</x>"
        ),
    )

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            payload,
            filename="batch.xlsx",
        )

    _assert_code(exc_info, "UNSAFE_XML")


@pytest.mark.parametrize(
    ("code", "configure_payload"),
    [
        (
            "TOO_MANY_XLSX_PARTS",
            lambda monkeypatch, payload: (
                monkeypatch.setattr(
                    batch_service,
                    "BATCH_MAX_ZIP_MEMBERS",
                    4,
                ),
                _with_extra_zip_member(
                    payload,
                    name="customXml/item1.xml",
                    body=b"<x/>",
                ),
            )[1],
        ),
        (
            "XLSX_EXPANDED_TOO_LARGE",
            lambda monkeypatch, payload: (
                monkeypatch.setattr(
                    batch_service,
                    "BATCH_MAX_UNCOMPRESSED_BYTES",
                    256,
                ),
                _with_extra_zip_member(
                    payload,
                    name="customXml/item1.bin",
                    body=b"x" * 180,
                ),
            )[1],
        ),
        (
            "XLSX_MEMBER_TOO_LARGE",
            lambda monkeypatch, payload: (
                monkeypatch.setattr(
                    batch_service,
                    "BATCH_MAX_MEMBER_BYTES",
                    64,
                ),
                _with_extra_zip_member(
                    payload,
                    name="customXml/item1.bin",
                    body=b"x" * 65,
                ),
            )[1],
        ),
        (
            "UNSAFE_XLSX_PATH",
            lambda monkeypatch, payload: _with_extra_zip_member(
                payload,
                name="../escape.xml",
                body=b"<x/>",
            ),
        ),
    ],
)
def test_zip_container_hard_limits_fail_closed(
    monkeypatch,
    code: str,
    configure_payload,
):
    payload = configure_payload(
        monkeypatch,
        _workbook_bytes(),
    )

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            payload,
            filename="batch.xlsx",
        )

    _assert_code(exc_info, code)


def test_row_count_is_bounded_before_business_processing():
    rows = [
        _valid_row(index)
        for index in range(1, BATCH_MAX_ROWS + 2)
    ]
    payload = _workbook_bytes(rows=rows)

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            payload,
            filename="batch.xlsx",
        )

    _assert_code(exc_info, "TOO_MANY_ROWS")


def test_dimensionless_valid_workbook_is_parsed_without_raw_type_errors():
    payload = _remove_sheet_dimension(
        _workbook_bytes(
            rows=[
                _valid_row(1),
                _valid_row(2),
            ]
        )
    )

    workbook = load_workbook(
        io.BytesIO(payload),
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        worksheet = workbook[BATCH_SHEET_NAME]
        assert worksheet.max_column is None
        assert worksheet.max_row is None
        assert worksheet.calculate_dimension(force=True) == "A1:H3"
        assert worksheet.max_column == len(BATCH_COLUMNAS)
        assert worksheet.max_row == 3
    finally:
        workbook.close()

    parsed = parsear_excel_lotes(
        payload,
        filename="dimensionless.xlsx",
    )

    assert parsed.sheet_name == BATCH_SHEET_NAME
    assert parsed.row_count == 2
    assert list(parsed.dataframe.columns) == BATCH_COLUMNAS
    assert list(
        parsed.dataframe["Identificador_Lote"]
    ) == [
        "RODAL-001",
        "RODAL-002",
    ]


def test_dimensionless_workbook_still_enforces_too_many_rows():
    payload = _remove_sheet_dimension(
        _workbook_bytes(
            rows=[
                _valid_row(index)
                for index in range(1, BATCH_MAX_ROWS + 2)
            ]
        )
    )

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            payload,
            filename="dimensionless.xlsx",
        )

    _assert_code(exc_info, "TOO_MANY_ROWS")


def test_too_many_sheets_are_rejected():
    workbook = Workbook()
    workbook.active.title = BATCH_SHEET_NAME
    workbook.active.append(BATCH_COLUMNAS)
    workbook.active.append(_valid_row(1))

    for index in range(2, 6):
        worksheet = workbook.create_sheet(
            title=f"Extra_{index}"
        )
        worksheet.append(BATCH_COLUMNAS)

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            buffer.getvalue(),
            filename="batch.xlsx",
        )

    _assert_code(exc_info, "TOO_MANY_SHEETS")


def test_header_formula_is_rejected():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = BATCH_SHEET_NAME
    worksheet.append(
        [
            "=A1",
            *BATCH_COLUMNAS[1:],
        ]
    )
    worksheet.append(_valid_row(1))

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            buffer.getvalue(),
            filename="batch.xlsx",
        )

    _assert_code(exc_info, "FORMULA_NOT_ALLOWED")


def test_lazy_malformed_worksheet_xml_fails_closed_without_raw_openpyxl_exception():
    payload = _rewrite_member(
        _workbook_bytes(
            rows=[
                _valid_row(1),
                _valid_row(2),
            ]
        ),
        name="xl/worksheets/sheet1.xml",
        transform=lambda body: body.replace(
            b"</sheetData>",
            b"",
            1,
        ),
    )

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            payload,
            filename="malformed.xlsx",
        )

    assert exc_info.value.code in {
        "INVALID_WORKBOOK",
        "INVALID_XLSX_STRUCTURE",
    }
    assert "Traceback" not in exc_info.value.detail


def test_blank_rows_are_ignored_without_inventing_records():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = BATCH_SHEET_NAME
    worksheet.append(BATCH_COLUMNAS)
    worksheet.append(_valid_row(1))
    worksheet.append([None] * len(BATCH_COLUMNAS))
    worksheet.append(_valid_row(2))

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    parsed = parsear_excel_lotes(
        buffer.getvalue(),
        filename="batch.xlsx",
    )

    assert parsed.row_count == 2
    assert list(
        parsed.dataframe["Identificador_Lote"]
    ) == [
        "RODAL-001",
        "RODAL-002",
    ]


def test_header_only_workbook_is_rejected():
    payload = _workbook_bytes(rows=[])

    with pytest.raises(
        BatchExcelValidationError
    ) as exc_info:
        parsear_excel_lotes(
            payload,
            filename="batch.xlsx",
        )

    _assert_code(exc_info, "NO_DATA_ROWS")
