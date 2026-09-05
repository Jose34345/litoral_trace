from __future__ import annotations

from io import BytesIO
import os
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, select, text

from litoral_trace.db.models import (
    AssuranceDocument,
    DocumentExtractionRun,
    ExtractedDocumentField,
    ReconciliationIssue,
    UsLaceyOperationField,
    VaultDocument,
)
from litoral_trace.db.tenant import set_tenant_db_context
from litoral_trace.us_lacey.access import UsLaceyOperationalAccessError
from litoral_trace.us_lacey.commercial import UsLaceyCommercialConfig
from litoral_trace.us_lacey.db import get_us_lacey_db_session, reset_us_lacey_engine_state
from litoral_trace.us_lacey.jobs import enqueue_us_lacey_document_job
from litoral_trace.us_lacey.operations import UsLaceyOperationService
from litoral_trace.us_lacey.projection import project_assurance_document_to_us_lacey
from litoral_trace.us_lacey.review import export_us_lacey_xlsx, review_us_lacey_field
from litoral_trace.us_lacey.self_service import register_us_lacey_company, verify_us_lacey_email
from litoral_trace.us_lacey.workflow import create_us_lacey_customer_operation


pytestmark = pytest.mark.skipif(
    os.environ.get("ENABLE_POSTGRES_TESTS") != "1"
    or not os.environ.get("US_LACEY_DATABASE_URL")
    or not os.environ.get("TEST_POSTGRES_MIGRATION_DATABASE_URL"),
    reason="requires the isolated U.S. PostgreSQL integration database",
)


def _commercial_config() -> UsLaceyCommercialConfig:
    return UsLaceyCommercialConfig(
        price_cents=12500,
        monthly_operation_limit=25,
        payment_provider="WISE",
        bank_transfer_instructions="CI-only payment instructions",
        terms_version="terms-workspace-v1",
        privacy_version="privacy-workspace-v1",
        beta_terms_version="beta-workspace-v1",
        support_email="support@litoraltrace.com",
    )


def _activate_account(organization_id: int) -> None:
    engine = create_engine(
        os.environ["TEST_POSTGRES_MIGRATION_DATABASE_URL"],
        pool_pre_ping=True,
        hide_parameters=True,
    )
    with engine.begin() as connection:
        connection.execute(
            text(
                "UPDATE public.us_lacey_organization_profiles "
                "SET account_status='ACTIVE', updated_at=now() "
                "WHERE organization_id=:organization_id"
            ),
            {"organization_id": organization_id},
        )
        connection.execute(
            text(
                "UPDATE public.us_lacey_subscriptions "
                "SET status='ACTIVE', started_at=coalesce(started_at, now()), updated_at=now() "
                "WHERE organization_id=:organization_id"
            ),
            {"organization_id": organization_id},
        )
        connection.execute(
            text(
                "UPDATE public.us_lacey_payments "
                "SET status='VERIFIED', verified_at=now(), paid_at=now(), updated_at=now() "
                "WHERE organization_id=:organization_id"
            ),
            {"organization_id": organization_id},
        )
    engine.dispose()


def _add_extracted_document(
    *,
    organization_id: int,
    user_id: int,
    filename: str,
    hts_value: str,
    country_of_harvest: str,
) -> int:
    session = get_us_lacey_db_session()
    try:
        set_tenant_db_context(session, organization_id)
        unique = uuid4().hex
        vault = VaultDocument(
            organization_id=organization_id,
            created_by_user_id=user_id,
            original_filename=filename,
            content_type="text/csv",
            size_bytes=128,
            sha256=(unique * 4)[:64],
            object_key=f"us-lacey/ci/tenants/{organization_id}/objects/{unique}",
            storage_backend="s3",
            storage_bucket="us-lacey-ci-private",
            document_type="OTHER_EVIDENCE",
            status="available",
        )
        session.add(vault)
        session.flush()
        assurance = AssuranceDocument(
            organization_id=organization_id,
            vault_document_id=vault.id,
            semantic_document_type="UNKNOWN",
            type_confidence=0.50,
            processing_status="NEEDS_REVIEW",
        )
        session.add(assurance)
        session.flush()
        run = DocumentExtractionRun(
            organization_id=organization_id,
            assurance_document_id=assurance.id,
            engine="assurance-deterministic-parser",
            engine_version="1.2.0",
            status="SUCCEEDED",
        )
        session.add(run)
        session.flush()
        rows = (
            ("raw.table.1.HTS Code", hts_value, "sheet:Shipment;data_row:1;column:1"),
            (
                "raw.table.1.Country of Harvest",
                country_of_harvest,
                "sheet:Shipment;data_row:1;column:2",
            ),
            (
                "raw.table.1.Merchandise Description",
                "White oak boards",
                "sheet:Shipment;data_row:1;column:3",
            ),
            (
                "raw.table.1.Species",
                "Quercus alba",
                "sheet:Shipment;data_row:1;column:4",
            ),
            # This ambiguous generic field must NOT become country_of_harvest.
            ("origin", "Canada", "sheet:Shipment;data_row:1;column:5"),
        )
        for field_name, value, locator in rows:
            session.add(
                ExtractedDocumentField(
                    organization_id=organization_id,
                    assurance_document_id=assurance.id,
                    extraction_run_id=run.id,
                    field_name=field_name,
                    original_value=value,
                    normalized_value=value,
                    value_type="cell",
                    confidence=0.98,
                    confidence_level="HIGH",
                    source_locator=locator,
                    auto_accepted=False,
                    needs_review=True,
                )
            )
        session.commit()
        return assurance.id
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def test_active_customer_operation_projection_review_export_and_conflict():
    reset_us_lacey_engine_state()
    suffix = uuid4().hex[:12]
    email = f"workspace-{suffix}@example.com"
    registered = register_us_lacey_company(
        legal_name=f"Workspace Imports {suffix} LLC",
        business_type="IMPORTER",
        admin_name="Workspace Admin",
        admin_email=email,
        password="correct-horse-workspace-123",
        commercial_config=_commercial_config(),
    )
    verify_us_lacey_email(registered.verification_token)

    with pytest.raises(UsLaceyOperationalAccessError):
        create_us_lacey_customer_operation(
            organization_id=registered.organization_id,
            user_id=registered.user_id,
            client_reference=f"SHIP-{suffix}",
        )

    _activate_account(registered.organization_id)
    operation = create_us_lacey_customer_operation(
        organization_id=registered.organization_id,
        user_id=registered.user_id,
        client_reference=f"SHIP-{suffix}",
        importer_name=f"Workspace Imports {suffix} LLC",
        line_references=("1",),
    )
    service = UsLaceyOperationService()
    operation_id = service.get_internal_id(
        organization_id=registered.organization_id,
        operation_public_id=operation.public_id,
    )

    owner = create_engine(
        os.environ["TEST_POSTGRES_MIGRATION_DATABASE_URL"],
        pool_pre_ping=True,
        hide_parameters=True,
    )
    with owner.connect() as connection:
        used = connection.execute(
            text(
                "SELECT used_operations FROM public.us_lacey_subscriptions "
                "WHERE organization_id=:organization_id"
            ),
            {"organization_id": registered.organization_id},
        ).scalar_one()
    owner.dispose()
    assert used == 1

    first_document_id = _add_extracted_document(
        organization_id=registered.organization_id,
        user_id=registered.user_id,
        filename="supplier-shipment.csv",
        hts_value="4407.91",
        country_of_harvest="United States",
    )
    service.attach_document(
        organization_id=registered.organization_id,
        operation_public_id=operation.public_id,
        assurance_document_id=first_document_id,
        document_role="SUPPLIER_DECLARATION",
    )
    queued = enqueue_us_lacey_document_job(
        organization_id=registered.organization_id,
        operation_id=operation_id,
        assurance_document_id=first_document_id,
    )
    queued_again = enqueue_us_lacey_document_job(
        organization_id=registered.organization_id,
        operation_id=operation_id,
        assurance_document_id=first_document_id,
    )
    assert queued.id == queued_again.id
    assert queued.status == "QUEUED"

    # The projection layer is independently deterministic from queue ownership.
    projection = project_assurance_document_to_us_lacey(
        organization_id=registered.organization_id,
        operation_id=operation_id,
        assurance_document_id=first_document_id,
    )
    assert projection.projected_count >= 5  # includes derived genus proposal

    detail = service.get_detail(
        organization_id=registered.organization_id,
        operation_public_id=operation.public_id,
    )
    by_name = {field.field_name: field for field in detail.fields}
    assert by_name["hts_code"].proposed_value == "4407.91"
    assert by_name["country_of_harvest"].proposed_value == "United States"
    assert by_name["merchandise_description"].proposed_value == "White oak boards"
    assert by_name["species"].proposed_value == "Quercus alba"
    assert by_name["genus"].proposed_value == "Quercus"
    assert by_name["country_of_harvest"].proposed_value != "Canada"
    assert by_name["hts_code"].status == "REVIEW"

    reviewed = review_us_lacey_field(
        organization_id=registered.organization_id,
        operation_public_id=operation.public_id,
        field_id=by_name["hts_code"].id,
        user_id=registered.user_id,
        user_email=email,
        action="accept",
    )
    assert reviewed.field_status == "MATCHED"

    workbook_bytes = export_us_lacey_xlsx(
        organization_id=registered.organization_id,
        operation_public_id=operation.public_id,
    )
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)
    assert {"Read Me", "Preparation Data", "Evidence", "Exceptions"}.issubset(
        set(workbook.sheetnames)
    )
    readme_values = [cell.value for row in workbook["Read Me"].iter_rows() for cell in row]
    assert any("not a legal compliance determination" in str(value) for value in readme_values)
    workbook.close()

    second_document_id = _add_extracted_document(
        organization_id=registered.organization_id,
        user_id=registered.user_id,
        filename="commercial-invoice.csv",
        hts_value="9999.99",
        country_of_harvest="United States",
    )
    service.attach_document(
        organization_id=registered.organization_id,
        operation_public_id=operation.public_id,
        assurance_document_id=second_document_id,
        document_role="COMMERCIAL_INVOICE",
    )
    second_projection = project_assurance_document_to_us_lacey(
        organization_id=registered.organization_id,
        operation_id=operation_id,
        assurance_document_id=second_document_id,
    )
    assert second_projection.conflict_count >= 1

    session = get_us_lacey_db_session()
    try:
        set_tenant_db_context(session, registered.organization_id)
        conflict = session.scalar(
            select(ReconciliationIssue).where(
                ReconciliationIssue.organization_id == registered.organization_id,
                ReconciliationIssue.operation_reference == f"us_lacey:{operation.public_id}",
                ReconciliationIssue.field_name == "hts_code",
                ReconciliationIssue.status == "OPEN",
            )
        )
        assert conflict is not None
        assert {conflict.left_value, conflict.right_value} == {"4407.91", "9999.99"}
        field = session.scalar(
            select(UsLaceyOperationField).where(
                UsLaceyOperationField.organization_id == registered.organization_id,
                UsLaceyOperationField.operation_id == operation_id,
                UsLaceyOperationField.field_name == "hts_code",
            )
        )
        assert field.field_status == "REVIEW"
        hts_field_id = field.id
    finally:
        session.rollback()
        session.close()

    resolved = review_us_lacey_field(
        organization_id=registered.organization_id,
        operation_public_id=operation.public_id,
        field_id=hts_field_id,
        user_id=registered.user_id,
        user_email=email,
        action="edit",
        value="4407.91",
    )
    assert resolved.field_status == "MATCHED"
    assert resolved.open_conflict_count == 0
