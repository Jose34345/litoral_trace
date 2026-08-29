"""Low-friction deterministic parsers used by Assurance v1.

The parser layer extracts what is actually present in the supplied file and
records source locations. It does not invent missing business facts. OCR is
attempted only when a PDF lacks useful digital text; failure remains fail-closed
and is surfaced as ``ocr_required`` for human review.
"""
from __future__ import annotations

from dataclasses import dataclass, field
import csv
from importlib.metadata import PackageNotFoundError, version as package_version
from io import BytesIO, StringIO
import os
from pathlib import Path, PurePath
import re
from tempfile import TemporaryDirectory
from typing import Any, Iterable
import zipfile

from openpyxl import load_workbook


class DocumentParseError(ValueError):
    """Raised when a supported document cannot be parsed safely."""


@dataclass(frozen=True, slots=True)
class SourceLocation:
    page: int | None = None
    sheet: str | None = None
    row: int | None = None
    column: int | None = None
    locator: str | None = None


@dataclass(frozen=True, slots=True)
class ParsedTable:
    name: str
    headers: tuple[str, ...]
    rows: tuple[dict[str, Any], ...]
    source: SourceLocation


@dataclass(frozen=True, slots=True)
class ParsedDocument:
    file_kind: str
    text: str = ""
    tables: tuple[ParsedTable, ...] = ()
    metadata: dict[str, Any] = field(default_factory=dict)
    ocr_required: bool = False


_OLE_XLS_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_MIN_USEFUL_PDF_TEXT_CHARS = 12
_MIN_USEFUL_PDF_ALPHA_CHARS = 4
_OCR_MAX_PAGES = 20
_OCR_RENDER_SCALE = 2.0
_OCR_TIMEOUT_SECONDS = 20
_OCR_TIMEOUT_ENV = "LT_ASSURANCE_OCR_TIMEOUT_SECONDS"
_OCR_TIMEOUT_MIN_SECONDS = 5
_OCR_TIMEOUT_MAX_SECONDS = 120
_OCR_LANGUAGES = ("spa", "eng")
_OCR_LANGUAGE = "+".join(_OCR_LANGUAGES)
_TOTAL_MARKERS = frozenset(
    {
        "total",
        "subtotal",
        "totales",
        "total general",
        "observaciones",
        "observacion",
    }
)


def _ocr_timeout_seconds() -> int:
    """Resolve a bounded per-page OCR timeout without changing production defaults."""
    raw = str(os.getenv(_OCR_TIMEOUT_ENV, "")).strip()
    if not raw:
        return _OCR_TIMEOUT_SECONDS
    try:
        value = int(raw)
    except ValueError:
        return _OCR_TIMEOUT_SECONDS
    return max(_OCR_TIMEOUT_MIN_SECONDS, min(_OCR_TIMEOUT_MAX_SECONDS, value))


def _clean_cell(value: Any) -> Any:
    if isinstance(value, str):
        cleaned = re.sub(r"\s+", " ", value).strip()
        return cleaned or None
    return value


def _header_label(value: Any, index: int) -> str:
    cleaned = _clean_cell(value)
    if cleaned is None:
        return f"column_{index + 1}"
    label = str(cleaned).strip()
    return label[:255] or f"column_{index + 1}"


def _row_nonempty_count(row: Iterable[Any]) -> int:
    return sum(_clean_cell(value) is not None for value in row)


def detect_header_row(rows: list[list[Any]], *, scan_limit: int = 25) -> int | None:
    """Choose the most plausible tabular header among early rows."""
    best_index: int | None = None
    best_score = -1.0
    for index, row in enumerate(rows[:scan_limit]):
        cleaned = [_clean_cell(value) for value in row]
        nonempty = [value for value in cleaned if value is not None]
        if len(nonempty) < 2:
            continue
        text_cells = sum(isinstance(value, str) for value in nonempty)
        unique_cells = len({str(value).casefold() for value in nonempty})
        score = len(nonempty) + (text_cells / len(nonempty)) + (unique_cells / len(nonempty))
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _is_decorative_or_total_row(values: list[Any]) -> bool:
    cleaned = [_clean_cell(value) for value in values]
    nonempty = [value for value in cleaned if value is not None]
    if not nonempty:
        return True
    first = str(nonempty[0]).strip().casefold()
    if first in _TOTAL_MARKERS or first.startswith("total ") or first.startswith("subtotal "):
        return True
    return False


def _records_from_rows(
    rows: list[list[Any]],
    *,
    header_index: int,
) -> tuple[tuple[str, ...], tuple[dict[str, Any], ...]]:
    raw_headers = rows[header_index]
    headers = tuple(_header_label(value, index) for index, value in enumerate(raw_headers))
    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        values = padded[: len(headers)]
        if _is_decorative_or_total_row(values):
            continue
        record = {
            header: _clean_cell(value)
            for header, value in zip(headers, values)
        }
        if any(value is not None for value in record.values()):
            records.append(record)
    return headers, tuple(records)


def validate_xlsx_bytes(content: bytes) -> None:
    try:
        with zipfile.ZipFile(BytesIO(content), "r") as archive:
            names = set(archive.namelist())
            required = {"[Content_Types].xml", "xl/workbook.xml", "_rels/.rels"}
            if not required.issubset(names):
                raise DocumentParseError("El archivo no corresponde a un XLSX valido.")
    except zipfile.BadZipFile as exc:
        raise DocumentParseError("El archivo XLSX esta corrupto.") from exc


def validate_xls_bytes(content: bytes) -> None:
    if not content.startswith(_OLE_XLS_SIGNATURE):
        raise DocumentParseError("El archivo no corresponde a un XLS binario valido.")


def parse_xlsx(content: bytes) -> ParsedDocument:
    validate_xlsx_bytes(content)
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise DocumentParseError("No se pudo abrir el XLSX.") from exc

    tables: list[ParsedTable] = []
    sheet_names: list[str] = []
    try:
        for worksheet in workbook.worksheets:
            sheet_names.append(worksheet.title)
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            header_index = detect_header_row(rows)
            if header_index is None:
                continue
            headers, records = _records_from_rows(rows, header_index=header_index)
            if not records:
                continue
            tables.append(
                ParsedTable(
                    name=worksheet.title,
                    headers=headers,
                    rows=records,
                    source=SourceLocation(
                        sheet=worksheet.title,
                        row=header_index + 1,
                        locator=f"sheet:{worksheet.title};header_row:{header_index + 1}",
                    ),
                )
            )
    finally:
        workbook.close()

    return ParsedDocument(
        file_kind="XLSX",
        tables=tuple(tables),
        metadata={"sheet_names": sheet_names, "sheet_count": len(sheet_names)},
    )


def parse_xls(content: bytes) -> ParsedDocument:
    """Parse legacy XLS using pandas/xlrd, keeping all useful sheets."""
    validate_xls_bytes(content)
    try:
        import pandas as pd
    except Exception as exc:  # pragma: no cover - dependency gate
        raise DocumentParseError("Pandas no esta disponible para leer XLS.") from exc

    try:
        excel = pd.ExcelFile(BytesIO(content), engine="xlrd")
    except Exception as exc:
        raise DocumentParseError("No se pudo abrir el XLS legado.") from exc

    tables: list[ParsedTable] = []
    try:
        for sheet_name in excel.sheet_names:
            frame = excel.parse(sheet_name=sheet_name, header=None, dtype=object)
            rows = frame.where(frame.notna(), None).values.tolist()
            header_index = detect_header_row(rows)
            if header_index is None:
                continue
            headers, records = _records_from_rows(rows, header_index=header_index)
            if not records:
                continue
            tables.append(
                ParsedTable(
                    name=str(sheet_name),
                    headers=headers,
                    rows=records,
                    source=SourceLocation(
                        sheet=str(sheet_name),
                        row=header_index + 1,
                        locator=f"sheet:{sheet_name};header_row:{header_index + 1}",
                    ),
                )
            )
    finally:
        excel.close()

    return ParsedDocument(
        file_kind="XLS",
        tables=tuple(tables),
        metadata={"sheet_names": list(excel.sheet_names), "sheet_count": len(excel.sheet_names)},
    )


def _decode_csv(content: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            decoded = content.decode(encoding)
        except UnicodeDecodeError:
            continue
        if "\x00" in decoded:
            continue
        return decoded, encoding
    raise DocumentParseError("No se pudo detectar una codificacion CSV soportada.")


def parse_csv(content: bytes) -> ParsedDocument:
    decoded, encoding = _decode_csv(content)
    sample = decoded[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","

    reader = csv.reader(StringIO(decoded), delimiter=delimiter)
    rows = [list(row) for row in reader]
    header_index = detect_header_row(rows)
    if header_index is None:
        raise DocumentParseError("El CSV no contiene una cabecera util.")
    headers, records = _records_from_rows(rows, header_index=header_index)
    table = ParsedTable(
        name="csv",
        headers=headers,
        rows=records,
        source=SourceLocation(
            row=header_index + 1,
            locator=f"csv:header_row:{header_index + 1}",
        ),
    )
    return ParsedDocument(
        file_kind="CSV",
        tables=(table,),
        metadata={
            "encoding": encoding,
            "delimiter": delimiter,
            "row_count": len(records),
        },
    )


def _has_useful_pdf_text(text: str) -> bool:
    compact = re.sub(r"\s+", "", text or "")
    alpha_chars = sum(character.isalpha() for character in compact)
    return (
        len(compact) >= _MIN_USEFUL_PDF_TEXT_CHARS
        and alpha_chars >= _MIN_USEFUL_PDF_ALPHA_CHARS
    )


def _safe_close(resource: object | None) -> None:
    close = getattr(resource, "close", None)
    if callable(close):
        try:
            close()
        except Exception:
            pass


def _extract_pdf_text_pages(content: bytes) -> tuple[str, int, int]:
    """Extract text from a PDF without logging or persisting document content."""
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(content), strict=False)
    page_texts: list[str] = []
    for page in reader.pages:
        try:
            page_texts.append((page.extract_text() or "").strip())
        except Exception:
            page_texts.append("")
    text = "\n\n".join(value for value in page_texts if value).strip()
    return text, len(reader.pages), sum(bool(value) for value in page_texts)


def _ocr_scanned_pdf_with_ocrmypdf(
    content: bytes,
    *,
    page_count: int,
) -> tuple[str, dict[str, Any]]:
    """Use OCRmyPDF as the primary OCR layer without mutating the Vault original."""
    timeout_seconds = _ocr_timeout_seconds()
    metadata: dict[str, Any] = {
        "ocr_attempted": True,
        "ocr_applied": False,
        "ocr_engine": "ocrmypdf",
        "ocr_language": _OCR_LANGUAGE,
        "ocr_pages_processed": 0,
        "ocr_timeout_seconds": timeout_seconds,
    }
    if page_count <= 0:
        metadata["ocr_error_code"] = "OCR_NO_PAGES"
        return "", metadata
    if page_count > _OCR_MAX_PAGES:
        metadata["ocr_error_code"] = "OCR_PAGE_LIMIT_EXCEEDED"
        metadata["ocr_page_limit"] = _OCR_MAX_PAGES
        return "", metadata

    try:
        import ocrmypdf
    except Exception as exc:
        metadata["ocr_error_code"] = "OCRMY_PDF_DEPENDENCY_UNAVAILABLE"
        metadata["ocr_error_type"] = type(exc).__name__
        return "", metadata

    try:
        metadata["ocr_engine_version"] = package_version("ocrmypdf")
    except PackageNotFoundError:
        metadata["ocr_engine_version"] = "unknown"

    try:
        with TemporaryDirectory(prefix="lt-assurance-ocr-") as temporary_directory:
            workdir = Path(temporary_directory)
            input_path = workdir / "input.pdf"
            output_path = workdir / "ocr.pdf"
            input_path.write_bytes(content)

            exit_code = ocrmypdf.ocr(
                input_path,
                output_path,
                language=list(_OCR_LANGUAGES),
                output_type="pdf",
                mode="skip",
                jobs=1,
                use_threads=True,
                optimize=0,
                tesseract_timeout=float(timeout_seconds),
                progress_bar=False,
            )
            metadata["ocr_exit_code"] = int(exit_code)
            output_content = output_path.read_bytes()
            text, output_page_count, pages_with_text = _extract_pdf_text_pages(output_content)
            metadata["ocr_pages_processed"] = output_page_count
            metadata["ocr_pages_with_text"] = pages_with_text
            metadata["ocr_output_size_bytes"] = len(output_content)
    except Exception as exc:
        metadata["ocr_error_code"] = "OCRMY_PDF_EXECUTION_FAILED"
        metadata["ocr_error_type"] = type(exc).__name__
        return "", metadata

    if not _has_useful_pdf_text(text):
        metadata["ocr_error_code"] = "OCRMY_PDF_NO_USEFUL_TEXT"
        return "", metadata

    metadata["ocr_applied"] = True
    return text, metadata


def _ocr_scanned_pdf_with_tesseract(
    content: bytes,
    *,
    page_count: int,
) -> tuple[str, dict[str, Any]]:
    """Legacy page-level Tesseract OCR retained only as a fail-closed fallback."""
    timeout_seconds = _ocr_timeout_seconds()
    metadata: dict[str, Any] = {
        "ocr_attempted": True,
        "ocr_applied": False,
        "ocr_engine": "tesseract",
        "ocr_language": _OCR_LANGUAGE,
        "ocr_pages_processed": 0,
        "ocr_timeout_seconds": timeout_seconds,
    }
    if page_count <= 0:
        metadata["ocr_error_code"] = "OCR_NO_PAGES"
        return "", metadata
    if page_count > _OCR_MAX_PAGES:
        metadata["ocr_error_code"] = "OCR_PAGE_LIMIT_EXCEEDED"
        metadata["ocr_page_limit"] = _OCR_MAX_PAGES
        return "", metadata

    try:
        import pypdfium2 as pdfium
        import pytesseract
    except Exception:
        metadata["ocr_error_code"] = "OCR_DEPENDENCY_UNAVAILABLE"
        return "", metadata

    document = None
    page_texts: list[str] = []
    try:
        document = pdfium.PdfDocument(content)
        for page_index in range(page_count):
            page = document[page_index]
            bitmap = None
            image = None
            try:
                bitmap = page.render(scale=_OCR_RENDER_SCALE)
                image = bitmap.to_pil().convert("L")
                extracted = pytesseract.image_to_string(
                    image,
                    lang=_OCR_LANGUAGE,
                    config="--psm 6",
                    timeout=timeout_seconds,
                )
                page_texts.append((extracted or "").strip())
                metadata["ocr_pages_processed"] = page_index + 1
            finally:
                _safe_close(image)
                _safe_close(bitmap)
                _safe_close(page)
    except Exception as exc:
        metadata["ocr_error_code"] = "OCR_EXECUTION_FAILED"
        metadata["ocr_error_type"] = type(exc).__name__
        return "", metadata
    finally:
        _safe_close(document)

    text = "\n\n".join(value for value in page_texts if value).strip()
    if not _has_useful_pdf_text(text):
        metadata["ocr_error_code"] = "OCR_NO_USEFUL_TEXT"
        return "", metadata

    metadata["ocr_applied"] = True
    metadata["ocr_pages_with_text"] = sum(bool(value) for value in page_texts)
    return text, metadata


def _ocr_scanned_pdf(content: bytes, *, page_count: int) -> tuple[str, dict[str, Any]]:
    """Prefer OCRmyPDF and fall back to the legacy page-level Tesseract path."""
    primary_text, primary_metadata = _ocr_scanned_pdf_with_ocrmypdf(
        content,
        page_count=page_count,
    )
    if _has_useful_pdf_text(primary_text):
        primary_metadata["ocr_fallback_used"] = False
        return primary_text, primary_metadata

    fallback_text, fallback_metadata = _ocr_scanned_pdf_with_tesseract(
        content,
        page_count=page_count,
    )
    merged_metadata = dict(fallback_metadata)
    merged_metadata["ocr_engine"] = "tesseract_fallback"
    merged_metadata["ocr_fallback_used"] = True
    merged_metadata["ocr_fallback_engine"] = "tesseract"
    merged_metadata["ocr_primary_engine"] = "ocrmypdf"
    if primary_metadata.get("ocr_engine_version"):
        merged_metadata["ocr_primary_engine_version"] = primary_metadata["ocr_engine_version"]
    if primary_metadata.get("ocr_error_code"):
        merged_metadata["ocr_primary_error_code"] = primary_metadata["ocr_error_code"]
    if primary_metadata.get("ocr_error_type"):
        merged_metadata["ocr_primary_error_type"] = primary_metadata["ocr_error_type"]
    if primary_metadata.get("ocr_exit_code") is not None:
        merged_metadata["ocr_primary_exit_code"] = primary_metadata["ocr_exit_code"]
    return fallback_text, merged_metadata


def parse_pdf(content: bytes) -> ParsedDocument:
    if not content.startswith(b"%PDF-"):
        raise DocumentParseError("El archivo no corresponde a un PDF valido.")
    if b"%%EOF" not in content[-4096:]:
        raise DocumentParseError("El PDF no contiene un cierre valido.")

    try:
        useful_text, page_count, pages_with_text = _extract_pdf_text_pages(content)
    except ImportError as exc:  # pragma: no cover - dependency gate
        raise DocumentParseError("pypdf no esta disponible.") from exc
    except Exception as exc:
        raise DocumentParseError("No se pudo abrir el PDF.") from exc

    ocr_required = not _has_useful_pdf_text(useful_text)
    ocr_metadata: dict[str, Any] = {
        "ocr_attempted": False,
        "ocr_applied": False,
    }
    if ocr_required:
        ocr_text, ocr_metadata = _ocr_scanned_pdf(content, page_count=page_count)
        if _has_useful_pdf_text(ocr_text):
            useful_text = ocr_text
            ocr_required = False

    tables: list[ParsedTable] = []
    if not ocr_required:
        try:
            import pdfplumber

            with pdfplumber.open(BytesIO(content)) as pdf:
                for page_number, page in enumerate(pdf.pages, start=1):
                    for table_index, raw_table in enumerate(page.extract_tables() or [], start=1):
                        rows = [list(row or []) for row in raw_table if row]
                        header_index = detect_header_row(rows, scan_limit=10)
                        if header_index is None:
                            continue
                        headers, records = _records_from_rows(rows, header_index=header_index)
                        if not records:
                            continue
                        tables.append(
                            ParsedTable(
                                name=f"page_{page_number}_table_{table_index}",
                                headers=headers,
                                rows=records,
                                source=SourceLocation(
                                    page=page_number,
                                    row=header_index + 1,
                                    locator=(
                                        f"pdf:page:{page_number};table:{table_index};"
                                        f"header_row:{header_index + 1}"
                                    ),
                                ),
                            )
                        )
        except Exception:
            # Table extraction is best-effort; extracted/OCR text remains authoritative.
            tables = []

    metadata = {
        "page_count": page_count,
        "pages_with_text": pages_with_text,
    }
    metadata.update(ocr_metadata)
    return ParsedDocument(
        file_kind="PDF",
        text=useful_text,
        tables=tuple(tables),
        metadata=metadata,
        ocr_required=ocr_required,
    )


def parse_document(filename: str, content: bytes) -> ParsedDocument:
    extension = PurePath(filename).suffix.lower()
    if extension == ".pdf":
        return parse_pdf(content)
    if extension == ".xlsx":
        return parse_xlsx(content)
    if extension == ".xls":
        return parse_xls(content)
    if extension == ".csv":
        return parse_csv(content)
    raise DocumentParseError("Formato no soportado: se acepta PDF, XLSX, XLS o CSV.")