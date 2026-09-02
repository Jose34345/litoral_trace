"""Human review and evidence-preserving exports for U.S. operations."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO, StringIO
import csv
from uuid import UUID

from openpyxl import Workbook
from sqlalchemy import func, select

from litoral_trace.db.models import (
    AssuranceDocument,
    ReconciliationIssue,
    UsLaceyFieldCandidate,
    UsLaceyOperation,
    UsLaceyOperationField,
    UsLaceyPpqPlantLine,
    UsLaceyPlantDeclaration,
    UsLaceyProcessingJob,
    VaultDocument,
)
from litoral_trace.db.tenant import set_tenant_db_context
from litoral_trace.services.audit import (
    AuditAction,
    AuditActor,
    AuditOutcome,
    record_audit_event,
)
from litoral_trace.us_lacey.db import get_us_lacey_db_session
from litoral_trace.us_lacey.domain import US_LACEY_REVIEW_FIELDS
from litoral_trace.us_lacey.operations import UsLaceyOperationNotFound
from litoral_trace.us_lacey.projection import refresh_us_lacey_operation_status
from litoral_trace.us_lacey.ppq505 import (
    PPQ505_FIELDS,
    PPQ505_PLANT_FIELDS,
    PPQ505_SHIPMENT_FIELDS,
    PPQ505_SHIPMENT_REFERENCE,
    is_paper_or_paperboard,
    not_required_allowed,
    validate_ppq_value,
)


class UsLaceyReviewError(RuntimeError):
    pass


@dataclass(frozen=True, slots=True)
class UsLaceyReviewResult:
    field_id: int
    field_status: str
    operation_status: str
    remaining_review_count: int
    remaining_missing_count: int
    open_conflict_count: int


@dataclass(frozen=True, slots=True)
class UsLaceyFinalizeResult:
    operation_status: str
    review_result: str


_LABELS = dict(US_LACEY_REVIEW_FIELDS)


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _operation(session, *, organization_id: int, operation_public_id: UUID | str):
    try:
        public_id = operation_public_id if isinstance(operation_public_id, UUID) else UUID(str(operation_public_id))
    except (ValueError, TypeError, AttributeError) as exc:
        raise UsLaceyOperationNotFound("Operation not found.") from exc
    operation = session.scalar(
        select(UsLaceyOperation).where(
            UsLaceyOperation.organization_id == organization_id,
            UsLaceyOperation.public_id == public_id,
        )
    )
    if operation is None:
        raise UsLaceyOperationNotFound("Operation not found.")
    return operation


def _counts(session, *, organization_id: int, operation: UsLaceyOperation) -> tuple[int, int, int]:
    review = session.scalar(
        select(func.count(UsLaceyOperationField.id)).where(
            UsLaceyOperationField.organization_id == organization_id,
            UsLaceyOperationField.operation_id == operation.id,
            UsLaceyOperationField.field_status == "REVIEW",
        )
    ) or 0
    missing = session.scalar(
        select(func.count(UsLaceyOperationField.id)).where(
            UsLaceyOperationField.organization_id == organization_id,
            UsLaceyOperationField.operation_id == operation.id,
            UsLaceyOperationField.field_status == "MISSING",
        )
    ) or 0
    conflicts = session.scalar(
        select(func.count(ReconciliationIssue.id)).where(
            ReconciliationIssue.organization_id == organization_id,
            ReconciliationIssue.operation_reference == f"us_lacey:{operation.public_id}",
            ReconciliationIssue.status == "OPEN",
        )
    ) or 0
    return int(review), int(missing), int(conflicts)


def review_us_lacey_field(
    *,
    organization_id: int,
    operation_public_id: UUID | str,
    field_id: int,
    user_id: int,
    user_email: str,
    action: str,
    value: str | None = None,
    candidate_id: int | None = None,
    reason_code: str | None = None,
) -> UsLaceyReviewResult:
    """Accept, edit or explicitly mark one preparation field as not required.

    Extracted originals/provenance are never overwritten. Human decisions live in
    ``human_value`` and reviewer metadata. Resolving a contradiction is explicit
    and audit logged.
    """
    org_id = int(organization_id)
    normalized_action = str(action or "").strip().lower()
    if normalized_action not in {"accept", "edit", "not_required"}:
        raise UsLaceyReviewError("Review action is invalid.")

    session = get_us_lacey_db_session()
    try:
        set_tenant_db_context(session, org_id)
        operation = _operation(
            session,
            organization_id=org_id,
            operation_public_id=operation_public_id,
        )
        field = session.scalar(
            select(UsLaceyOperationField).where(
                UsLaceyOperationField.organization_id == org_id,
                UsLaceyOperationField.operation_id == operation.id,
                UsLaceyOperationField.id == int(field_id),
            )
        )
        if field is None:
            raise UsLaceyReviewError("Review field was not found for this operation.")

        selected_candidate = None
        if candidate_id is not None:
            selected_candidate = session.scalar(
                select(UsLaceyFieldCandidate).where(
                    UsLaceyFieldCandidate.organization_id == org_id,
                    UsLaceyFieldCandidate.operation_id == operation.id,
                    UsLaceyFieldCandidate.operation_field_id == field.id,
                    UsLaceyFieldCandidate.id == int(candidate_id),
                )
            )
            if selected_candidate is None:
                raise UsLaceyReviewError("The selected source candidate was not found.")
        proposed = (
            (selected_candidate.normalized_value or selected_candidate.original_value)
            if selected_candidate is not None
            else (field.normalized_value or field.original_value)
        )
        before = {
            "field_id": field.id,
            "field_name": field.field_name,
            "field_status": field.field_status,
            "effective_value": field.human_value or proposed,
        }
        if normalized_action == "accept":
            if proposed is None or not str(proposed).strip():
                raise UsLaceyReviewError("There is no extracted value to accept.")
            validation = validate_ppq_value(field.field_name, proposed)
            if validation.status.value in {"INVALID", "MISSING", "REVIEW_REQUIRED"}:
                raise UsLaceyReviewError(validation.error or "This value requires further review.")
            field.human_value = validation.normalized_value
            field.field_status = "MATCHED"
            field.validation_status = "VALID"
            field.validation_error = None
            field.not_required_reason_code = None
        elif normalized_action == "edit":
            human = str(value or "").strip()
            if not human:
                raise UsLaceyReviewError("Enter a value before saving this review.")
            if len(human) > 4000:
                raise UsLaceyReviewError("Reviewed value is too long.")
            validation = validate_ppq_value(field.field_name, human)
            if validation.status.value in {"INVALID", "MISSING", "REVIEW_REQUIRED"}:
                raise UsLaceyReviewError(validation.error or "This value requires further review.")
            field.human_value = validation.normalized_value
            field.field_status = "MATCHED"
            field.validation_status = "VALID"
            field.validation_error = None
            field.not_required_reason_code = None
        else:
            normalized_reason = str(reason_code or "").strip().upper()
            context_field = session.scalar(
                select(UsLaceyOperationField).where(
                    UsLaceyOperationField.organization_id == org_id,
                    UsLaceyOperationField.operation_id == operation.id,
                    UsLaceyOperationField.merchandise_line_reference == field.merchandise_line_reference,
                    UsLaceyOperationField.field_name == "article_component",
                )
            )
            article_or_product = (
                None if context_field is None
                else (context_field.human_value or context_field.normalized_value or context_field.original_value)
            )
            if not not_required_allowed(
                field.field_name, normalized_reason, article_or_product=article_or_product
            ):
                raise UsLaceyReviewError(
                    "NOT_REQUIRED is not allowed for this PPQ field and context."
                )
            field.human_value = None
            field.field_status = "NOT_REQUIRED"
            field.validation_status = "VALID"
            field.validation_error = None
            field.not_required_reason_code = normalized_reason

        field.reviewed_by_user_id = int(user_id)
        field.reviewed_at = _utc_now()

        if selected_candidate is not None:
            # The selected evidence, not a previously displayed candidate,
            # becomes the final field provenance.
            field.original_value = selected_candidate.original_value
            field.normalized_value = selected_candidate.normalized_value
            field.source_assurance_document_id = selected_candidate.source_assurance_document_id
            field.source_page = selected_candidate.source_page
            field.source_locator = selected_candidate.source_locator
            field.extractor = selected_candidate.extractor
            field.extractor_version = selected_candidate.extractor_version
            field.confidence = selected_candidate.confidence
            related_candidates = session.scalars(
                select(UsLaceyFieldCandidate).where(
                    UsLaceyFieldCandidate.organization_id == org_id,
                    UsLaceyFieldCandidate.operation_field_id == field.id,
                )
            ).all()
            for candidate in related_candidates:
                candidate.decision = "SELECTED" if candidate.id == selected_candidate.id else "REJECTED"
                candidate.decided_by_user_id = int(user_id)
                candidate.decided_at = field.reviewed_at

        open_issues = session.scalars(
            select(ReconciliationIssue).where(
                ReconciliationIssue.organization_id == org_id,
                ReconciliationIssue.operation_reference == f"us_lacey:{operation.public_id}",
                ReconciliationIssue.us_lacey_operation_field_id == field.id,
                ReconciliationIssue.status == "OPEN",
            )
        ).all()
        for issue in open_issues:
            issue.status = "RESOLVED"
            issue.resolution_justification = (
                f"Explicit human review ({normalized_action}) by U.S. portal user {user_id}."
            )
            issue.resolved_at = field.reviewed_at

        operation_status = refresh_us_lacey_operation_status(
            session,
            organization_id=org_id,
            operation=operation,
        )
        review_count, missing_count, conflict_count = _counts(
            session,
            organization_id=org_id,
            operation=operation,
        )
        actor = AuditActor(
            organization_id=org_id,
            user_id=int(user_id),
            username=str(user_email or "").strip() or None,
            role="us_lacey_customer",
        )
        record_audit_event(
            session,
            actor=actor,
            action=(
                AuditAction.ASSURANCE_REVIEW_APPROVE
                if normalized_action == "accept"
                else AuditAction.ASSURANCE_REVIEW_CORRECT
            ),
            entity_type="us_lacey_operation",
            entity_id=operation.id,
            outcome=AuditOutcome.SUCCESS,
            metadata={
                "operation_public_id": str(operation.public_id),
                "field_id": field.id,
                "field_name": field.field_name,
                "review_action": normalized_action,
                "resolved_conflict_count": len(open_issues),
            },
            before_data=before,
            after_data={
                "field_status": field.field_status,
                "effective_value": field.human_value or proposed,
                "operation_status": operation_status,
            },
            detail="U.S. document-preparation field reviewed by a customer user.",
        )
        session.commit()
        return UsLaceyReviewResult(
            field_id=field.id,
            field_status=field.field_status,
            operation_status=operation_status,
            remaining_review_count=review_count,
            remaining_missing_count=missing_count,
            open_conflict_count=conflict_count,
        )
    except (UsLaceyReviewError, UsLaceyOperationNotFound):
        session.rollback()
        raise
    except Exception as exc:
        session.rollback()
        raise UsLaceyReviewError("Unable to save this review decision.") from exc
    finally:
        session.close()


def finalize_us_lacey_review(
    *,
    organization_id: int,
    operation_public_id: UUID | str,
    user_id: int,
    user_email: str,
) -> UsLaceyFinalizeResult:
    """Close human review only when no unresolved preparation item remains."""
    org_id = int(organization_id)
    session = get_us_lacey_db_session()
    try:
        set_tenant_db_context(session, org_id)
        operation = _operation(
            session,
            organization_id=org_id,
            operation_public_id=operation_public_id,
        )
        active_jobs = session.scalar(
            select(func.count(UsLaceyProcessingJob.id)).where(
                UsLaceyProcessingJob.organization_id == org_id,
                UsLaceyProcessingJob.operation_id == operation.id,
                UsLaceyProcessingJob.status.in_(("QUEUED", "RUNNING", "RETRY")),
            )
        ) or 0
        failed_jobs = session.scalar(
            select(func.count(UsLaceyProcessingJob.id)).where(
                UsLaceyProcessingJob.organization_id == org_id,
                UsLaceyProcessingJob.operation_id == operation.id,
                UsLaceyProcessingJob.status == "FAILED",
            )
        ) or 0
        review_count, missing_count, conflict_count = _counts(
            session,
            organization_id=org_id,
            operation=operation,
        )
        plant_line_count = session.scalar(
            select(func.count(UsLaceyPlantDeclaration.id))
            .join(UsLaceyPpqPlantLine, UsLaceyPpqPlantLine.id == UsLaceyPlantDeclaration.plant_line_id)
            .where(
                UsLaceyPlantDeclaration.organization_id == org_id,
                UsLaceyPpqPlantLine.operation_id == operation.id,
            )
        ) or 0
        if int(operation.document_count) <= 0:
            raise UsLaceyReviewError("Upload at least one document before completing review.")
        if int(active_jobs):
            raise UsLaceyReviewError("Document processing is still in progress.")
        if int(failed_jobs):
            raise UsLaceyReviewError("A document-processing failure must be resolved first.")
        if int(plant_line_count) <= 0:
            raise UsLaceyReviewError(
                "At least one explicit plant declaration line is required before completion."
            )
        if review_count or missing_count or conflict_count:
            raise UsLaceyReviewError(
                "Resolve every missing field, review item and contradiction before completing review."
            )

        operation.status = "COMPLETED"
        operation.review_result = "DOCUMENT_REVIEW_COMPLETE"
        actor = AuditActor(
            organization_id=org_id,
            user_id=int(user_id),
            username=str(user_email or "").strip() or None,
            role="us_lacey_customer",
        )
        record_audit_event(
            session,
            actor=actor,
            action=AuditAction.ASSURANCE_REVIEW_APPROVE,
            entity_type="us_lacey_operation",
            entity_id=operation.id,
            outcome=AuditOutcome.SUCCESS,
            metadata={
                "operation_public_id": str(operation.public_id),
                "review_result": operation.review_result,
            },
            after_data={
                "operation_status": operation.status,
                "review_result": operation.review_result,
            },
            detail=(
                "Human document review completed. This is not a legal compliance determination."
            ),
        )
        session.commit()
        return UsLaceyFinalizeResult(
            operation_status=operation.status,
            review_result=operation.review_result,
        )
    except (UsLaceyReviewError, UsLaceyOperationNotFound):
        session.rollback()
        raise
    except Exception as exc:
        session.rollback()
        raise UsLaceyReviewError("Unable to complete operation review.") from exc
    finally:
        session.close()


def _export_rows(*, organization_id: int, operation_public_id: UUID | str):
    org_id = int(organization_id)
    session = get_us_lacey_db_session()
    try:
        set_tenant_db_context(session, org_id)
        operation = _operation(
            session,
            organization_id=org_id,
            operation_public_id=operation_public_id,
        )
        fields = session.scalars(
            select(UsLaceyOperationField)
            .where(
                UsLaceyOperationField.organization_id == org_id,
                UsLaceyOperationField.operation_id == operation.id,
            )
            .order_by(
                UsLaceyOperationField.merchandise_line_reference.asc(),
                UsLaceyOperationField.id.asc(),
            )
        ).all()
        candidates = session.scalars(
            select(UsLaceyFieldCandidate).where(
                UsLaceyFieldCandidate.organization_id == org_id,
                UsLaceyFieldCandidate.operation_id == operation.id,
                UsLaceyFieldCandidate.decision != "PENDING",
            ).order_by(UsLaceyFieldCandidate.id.asc())
        ).all()
        documents = session.scalars(
            select(AssuranceDocument).where(
                AssuranceDocument.organization_id == org_id,
                AssuranceDocument.id.in_(
                    tuple(
                        sorted(
                            {
                                field.source_assurance_document_id
                                for field in fields
                                if field.source_assurance_document_id is not None
                            } | {
                                candidate.source_assurance_document_id
                                for candidate in candidates
                            }
                        )
                    )
                    or (-1,)
                ),
            )
        ).all()
        filenames: dict[int, str] = {}
        for document in documents:
            vault = session.scalar(
                select(VaultDocument).where(
                    VaultDocument.organization_id == org_id,
                    VaultDocument.id == document.vault_document_id,
                )
            )
            if vault is not None:
                filenames[document.id] = vault.original_filename
        issues = session.scalars(
            select(ReconciliationIssue)
            .where(
                ReconciliationIssue.organization_id == org_id,
                ReconciliationIssue.operation_reference == f"us_lacey:{operation.public_id}",
            )
            .order_by(ReconciliationIssue.created_at.asc(), ReconciliationIssue.id.asc())
        ).all()
        return operation, tuple(fields), filenames, tuple(issues), tuple(candidates)
    finally:
        session.close()


def export_us_lacey_csv(
    *, organization_id: int, operation_public_id: UUID | str
) -> bytes:
    operation, fields, _filenames, _issues, _candidates = _export_rows(
        organization_id=organization_id,
        operation_public_id=operation_public_id,
    )
    line_refs = tuple(dict.fromkeys(
        field.merchandise_line_reference for field in fields if field.field_scope == "PLANT_LINE"
    ))
    by_key = {(field.merchandise_line_reference, field.field_name): field for field in fields}
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(["Plant Line Reference", *[f"PPQ #{field.number} {field.label}" for field in PPQ505_FIELDS]])
    for line in line_refs or ("",):
        values = []
        for contract in PPQ505_FIELDS:
            reference = PPQ505_SHIPMENT_REFERENCE if contract.scope.value == "SHIPMENT" else line
            field = by_key.get((reference, contract.key))
            values.append("" if field is None else (field.human_value or field.normalized_value or field.original_value or ""))
        writer.writerow([line, *values])
    return output.getvalue().encode("utf-8-sig")


def export_us_lacey_xlsx(
    *, organization_id: int, operation_public_id: UUID | str
) -> bytes:
    operation, fields, filenames, issues, candidates = _export_rows(
        organization_id=organization_id,
        operation_public_id=operation_public_id,
    )
    line_refs = tuple(dict.fromkeys(
        field.merchandise_line_reference for field in fields if field.field_scope == "PLANT_LINE"
    ))
    by_key = {(field.merchandise_line_reference, field.field_name): field for field in fields}

    workbook = Workbook()
    readme = workbook.active
    readme.title = "Read Me"
    readme.append(["Litoral Trace — PPQ Form 505 preparation workbook"])
    readme.append(["Operation", operation.client_reference])
    readme.append(["Workspace status", operation.status])
    readme.append(["Classification", "PREPARATION WORK PRODUCT"])
    readme.append(["Filing status", "NOT FILED"])
    readme.append(["Review requirement", "HUMAN REVIEW REQUIRED"])
    readme.append(["Submission status", "NO ACE/LAWGS SUBMISSION"])
    readme.append(["Important", "This workbook organizes source evidence. It is not a legal compliance determination, certification, government filing, or government acceptance."])
    readme.append(["Signature", "No preparer signature is generated or simulated."])

    shipment_sheet = workbook.create_sheet("Shipment Summary")
    shipment_sheet.append(["PPQ Number", "Field", "Value", "Review Status", "Validation"])
    for contract in PPQ505_SHIPMENT_FIELDS:
        field = by_key.get((PPQ505_SHIPMENT_REFERENCE, contract.key))
        shipment_sheet.append([
            contract.number, contract.label,
            "" if field is None else (field.human_value or field.normalized_value or field.original_value or ""),
            "MISSING" if field is None else field.field_status,
            "MISSING" if field is None else field.validation_status,
        ])
    shipment_sheet.freeze_panes = "A2"

    contract_sheet = workbook.create_sheet("PPQ 505 Fields")
    contract_sheet.append(["PPQ Number", "Scope", "Plant Line", "Field", "Value", "Review Status", "Validation", "Validation Error"])
    for contract in PPQ505_FIELDS:
        references = (PPQ505_SHIPMENT_REFERENCE,) if contract.scope.value == "SHIPMENT" else line_refs
        for reference in references:
            field = by_key.get((reference, contract.key))
            contract_sheet.append([
                contract.number, contract.scope.value,
                "" if reference == PPQ505_SHIPMENT_REFERENCE else reference,
                contract.label,
                "" if field is None else (field.human_value or field.normalized_value or field.original_value or ""),
                "MISSING" if field is None else field.field_status,
                "MISSING" if field is None else field.validation_status,
                "" if field is None else (field.validation_error or ""),
            ])
    contract_sheet.freeze_panes = "A2"
    contract_sheet.auto_filter.ref = contract_sheet.dimensions

    plant_sheet = workbook.create_sheet("Plant Lines")
    plant_sheet.append(["Plant Line Reference", *[f"PPQ #{field.number} {field.label}" for field in PPQ505_PLANT_FIELDS]])
    for line in line_refs:
        row = [line]
        for contract in PPQ505_PLANT_FIELDS:
            field = by_key.get((line, contract.key))
            row.append("" if field is None else (field.human_value or field.normalized_value or field.original_value or ""))
        plant_sheet.append(row)
    plant_sheet.freeze_panes = "A2"
    plant_sheet.auto_filter.ref = plant_sheet.dimensions

    evidence_sheet = workbook.create_sheet("Evidence")
    evidence_sheet.append(
        [
            "Scope", "Line Reference", "PPQ Number", "Field",
            "Original Source Value", "Normalized / Human Value",
            "Value",
            "Review Status",
            "Confidence",
            "Source Document",
            "Source Page",
            "Source Locator",
            "Extractor",
            "Extractor Version",
            "Reviewed By User ID",
            "Reviewed At",
        ]
    )
    for field in fields:
        evidence_sheet.append(
            [
                field.field_scope,
                "" if field.field_scope == "SHIPMENT" else field.merchandise_line_reference,
                next((item.number for item in PPQ505_FIELDS if item.key == field.field_name), ""),
                _LABELS.get(field.field_name, field.field_name),
                field.original_value or "",
                field.human_value or field.normalized_value or "",
                field.human_value or field.normalized_value or field.original_value or "",
                field.field_status,
                float(field.confidence),
                filenames.get(field.source_assurance_document_id or -1, ""),
                field.source_page,
                field.source_locator or "",
                field.extractor or "",
                field.extractor_version or "",
                field.reviewed_by_user_id,
                field.reviewed_at.isoformat() if field.reviewed_at else "",
            ]
        )
    evidence_sheet.freeze_panes = "A2"
    evidence_sheet.auto_filter.ref = evidence_sheet.dimensions

    exception_sheet = workbook.create_sheet("Exceptions")
    exception_sheet.append(
        [
            "Status",
            "Severity",
            "Field",
            "Left Value",
            "Right Value",
            "Explanation",
            "Resolution",
            "Reviewed / Resolved At",
        ]
    )
    for issue in issues:
        exception_sheet.append(
            [
                issue.status,
                issue.severity,
                issue.field_name or "",
                issue.left_value or "",
                issue.right_value or "",
                issue.explanation,
                issue.resolution_justification or "",
                issue.resolved_at.isoformat() if issue.resolved_at else "",
            ]
        )
    for candidate in candidates:
        exception_sheet.append(
            [
                candidate.decision,
                "HUMAN_DECISION",
                f"Candidate #{candidate.id}",
                candidate.original_value,
                candidate.normalized_value or "",
                f"Source document: {filenames.get(candidate.source_assurance_document_id, '')}",
                "Selected/rejected candidate",
                candidate.decided_at.isoformat() if candidate.decided_at else "",
            ]
        )
    for field in fields:
        if field.reviewed_at is None or field.field_status not in {"MATCHED", "NOT_REQUIRED"}:
            continue
        exception_sheet.append(
            [
                field.field_status,
                "HUMAN_DECISION",
                _LABELS.get(field.field_name, field.field_name),
                field.original_value or "",
                field.human_value or "",
                "Manual review" if field.human_value else "Contextual applicability",
                field.not_required_reason_code or "Manual override / accepted value",
                field.reviewed_at.isoformat(),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
