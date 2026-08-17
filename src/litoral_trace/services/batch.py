"""Secure XLSX ingestion and semantic row validation for Litoral Trace."""

from __future__ import annotations

from dataclasses import dataclass
import hashlib
import io
import math
from numbers import Real
from pathlib import PurePosixPath
import re
import unicodedata
import zipfile

from openpyxl import load_workbook
import pandas as pd

from litoral_trace.services.compliance import (
    evaluar_compliance_lote,
    generar_dds_json_traces_nt,
)
from litoral_trace.services.reports import generar_pdf_reporte_bytes


BATCH_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
BATCH_SHEET_NAME = "Plantilla_LitoralTrace"

BATCH_COLUMNAS = [
    "Identificador_Lote",
    "ID_Proveedor",
    "Producto_Forestal",
    "Hectareas",
    "Latitud",
    "Longitud",
    "Volumen_Ingresado_Ton",
    "Volumen_Exportar_Ton",
]

BATCH_FILA_EJEMPLO = [
    "Rodal_Norte_01",
    "CUIT-30123456789",
    "Madera Aserrada (Eucalipto)",
    120.0,
    -27.50,
    -58.90,
    500.0,
    200.0,
]

BATCH_MAX_FILE_BYTES = 10 * 1024 * 1024
BATCH_MAX_ROWS = 500
BATCH_MAX_SHEETS = 4
BATCH_MAX_ZIP_MEMBERS = 2048
BATCH_MAX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
BATCH_MAX_MEMBER_BYTES = 16 * 1024 * 1024
BATCH_MAX_TEXT_LENGTH = 100

_REQUIRED_XLSX_MEMBERS = frozenset(
    {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
)

_FORBIDDEN_MEMBER_PREFIXES = (
    "xl/embeddings/",
    "xl/externallinks/",
    "xl/activex/",
    "xl/ctrlprops/",
    "customui/",
)

_FORBIDDEN_MEMBER_FRAGMENTS = (
    "vbaproject.bin",
    "oleobject",
)

_CONTROL_CHARACTER_RE = re.compile(r"[\x00-\x1f\x7f]")


class BatchExcelError(ValueError):
    """Base class for safe, user-originated XLSX ingestion errors."""


class BatchExcelValidationError(BatchExcelError):
    """Fail-closed XLSX structural validation error with a stable public code."""

    def __init__(self, code: str, detail: str) -> None:
        super().__init__(detail)
        self.code = code
        self.detail = detail


class BatchSemanticValidationError(BatchExcelError):
    """Raised when one or more rows fail semantic validation."""

    def __init__(self, result: "BatchValidationResult") -> None:
        self.result = result
        super().__init__(
            (
                "La planilla contiene "
                f"{result.invalid_rows} fila(s) con errores de validación."
            )
        )


@dataclass(frozen=True)
class BatchWorkbook:
    """Canonical, structurally validated workbook payload."""

    filename: str
    sha256: str
    sheet_name: str
    row_count: int
    dataframe: pd.DataFrame
    source_row_numbers: tuple[int, ...]


@dataclass(frozen=True)
class BatchRowError:
    """Safe validation error suitable for later API/preview serialization."""

    row: int
    field: str
    code: str
    message: str


@dataclass(frozen=True)
class BatchCanonicalRow:
    """Fully normalized row ready for later persistence in P2.4C."""

    source_row: int
    identificador: str
    productor_id: str
    producto_forestal: str
    hectareas: float
    latitud: float
    longitud: float
    volumen_ingresado_ton: float
    volumen_exportar_ton: float

    def as_lote_payload(self) -> dict[str, object]:
        """Return the canonical persistence payload without tenant identity."""

        return {
            "identificador": self.identificador,
            "productor_id": self.productor_id,
            "producto_forestal": self.producto_forestal,
            "hectareas": self.hectareas,
            "latitud": self.latitud,
            "longitud": self.longitud,
            "volumen_ingresado_ton": self.volumen_ingresado_ton,
            "volumen_exportar_ton": self.volumen_exportar_ton,
        }


@dataclass(frozen=True)
class BatchRowValidation:
    """Validation outcome for a single source spreadsheet row."""

    row: int
    valid: bool
    data: BatchCanonicalRow | None
    errors: tuple[BatchRowError, ...]


@dataclass(frozen=True)
class BatchValidationResult:
    """Workbook-wide semantic validation result."""

    valid: bool
    total_rows: int
    valid_rows: int
    invalid_rows: int
    rows: tuple[BatchRowValidation, ...]

    @property
    def errors(self) -> tuple[BatchRowError, ...]:
        return tuple(
            error
            for row_result in self.rows
            for error in row_result.errors
        )

    @property
    def canonical_rows(self) -> tuple[BatchCanonicalRow, ...]:
        return tuple(
            row_result.data
            for row_result in self.rows
            if row_result.valid and row_result.data is not None
        )


def _raise_batch_error(code: str, detail: str) -> None:
    raise BatchExcelValidationError(code=code, detail=detail)


def normalizar_nombre_archivo_batch(filename: str | None) -> str:
    """Return a safe basename and require the modern XLSX container format."""

    raw = unicodedata.normalize(
        "NFKC",
        str(filename or ""),
    )
    raw = raw.replace("\\", "/")
    basename = PurePosixPath(raw).name.strip()

    if not basename:
        _raise_batch_error(
            "INVALID_FILENAME",
            "El archivo debe tener un nombre válido.",
        )

    if _CONTROL_CHARACTER_RE.search(basename):
        _raise_batch_error(
            "INVALID_FILENAME",
            "El nombre del archivo contiene caracteres no permitidos.",
        )

    if len(basename) > 255:
        _raise_batch_error(
            "INVALID_FILENAME",
            "El nombre del archivo excede el límite permitido.",
        )

    if not basename.lower().endswith(".xlsx"):
        _raise_batch_error(
            "UNSUPPORTED_FILE_TYPE",
            "Solo se admiten planillas Excel en formato .xlsx.",
        )

    return basename


def _validate_zip_member_name(name: str) -> None:
    normalized = name.replace("\\", "/")
    path = PurePosixPath(normalized)

    if (
        not normalized
        or normalized.startswith("/")
        or any(part in {"", ".", ".."} for part in path.parts)
        or ":" in path.parts[0]
    ):
        _raise_batch_error(
            "UNSAFE_XLSX_PATH",
            "La estructura interna del archivo XLSX no es válida.",
        )


def _read_member_bounded(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
) -> bytes:
    if info.file_size > BATCH_MAX_MEMBER_BYTES:
        _raise_batch_error(
            "XLSX_MEMBER_TOO_LARGE",
            "La planilla contiene una parte interna demasiado grande.",
        )

    with archive.open(info, "r") as member:
        payload = member.read(BATCH_MAX_MEMBER_BYTES + 1)

    if len(payload) > BATCH_MAX_MEMBER_BYTES:
        _raise_batch_error(
            "XLSX_MEMBER_TOO_LARGE",
            "La planilla contiene una parte interna demasiado grande.",
        )

    return payload


def _preflight_xlsx_container(payload: bytes) -> None:
    """Validate the ZIP/XML container before handing bytes to openpyxl."""

    if not payload.startswith(b"PK"):
        _raise_batch_error(
            "INVALID_XLSX_CONTAINER",
            "El archivo no contiene una estructura XLSX válida.",
        )

    try:
        archive = zipfile.ZipFile(
            io.BytesIO(payload),
            mode="r",
            allowZip64=False,
        )
    except (zipfile.BadZipFile, zipfile.LargeZipFile):
        _raise_batch_error(
            "INVALID_XLSX_CONTAINER",
            "El archivo no contiene una estructura XLSX válida.",
        )

    with archive:
        infos = [
            info
            for info in archive.infolist()
            if not info.is_dir()
        ]

        if not infos:
            _raise_batch_error(
                "INVALID_XLSX_CONTAINER",
                "El archivo XLSX no contiene datos internos.",
            )

        if len(infos) > BATCH_MAX_ZIP_MEMBERS:
            _raise_batch_error(
                "TOO_MANY_XLSX_PARTS",
                "La planilla excede la complejidad interna permitida.",
            )

        names: set[str] = set()
        total_uncompressed = 0

        for info in infos:
            _validate_zip_member_name(info.filename)

            normalized_name = info.filename.replace("\\", "/")
            lowered_name = normalized_name.lower()

            if normalized_name in names:
                _raise_batch_error(
                    "DUPLICATE_XLSX_PART",
                    "La estructura interna del archivo XLSX contiene duplicados.",
                )
            names.add(normalized_name)

            if info.flag_bits & 0x1:
                _raise_batch_error(
                    "ENCRYPTED_XLSX_PART",
                    "No se admiten planillas XLSX con partes cifradas.",
                )

            if any(
                lowered_name.startswith(prefix)
                for prefix in _FORBIDDEN_MEMBER_PREFIXES
            ) or any(
                fragment in lowered_name
                for fragment in _FORBIDDEN_MEMBER_FRAGMENTS
            ):
                _raise_batch_error(
                    "UNSUPPORTED_XLSX_FEATURE",
                    (
                        "La planilla contiene elementos embebidos, vínculos "
                        "externos o automatizaciones no permitidas."
                    ),
                )

            if info.file_size > BATCH_MAX_MEMBER_BYTES:
                _raise_batch_error(
                    "XLSX_MEMBER_TOO_LARGE",
                    "La planilla contiene una parte interna demasiado grande.",
                )

            total_uncompressed += info.file_size
            if total_uncompressed > BATCH_MAX_UNCOMPRESSED_BYTES:
                _raise_batch_error(
                    "XLSX_EXPANDED_TOO_LARGE",
                    "La planilla excede el tamaño interno permitido.",
                )

            if lowered_name.endswith((".xml", ".rels")):
                xml_payload = _read_member_bounded(
                    archive,
                    info,
                )
                upper_xml = xml_payload.upper()

                if b"<!DOCTYPE" in upper_xml or b"<!ENTITY" in upper_xml:
                    _raise_batch_error(
                        "UNSAFE_XML",
                        "La planilla contiene XML no permitido.",
                    )

                if lowered_name.endswith(".rels") and (
                    b'TARGETMODE="EXTERNAL"' in upper_xml
                    or b"TARGETMODE='EXTERNAL'" in upper_xml
                ):
                    _raise_batch_error(
                        "EXTERNAL_RELATIONSHIP",
                        "No se admiten vínculos externos en la planilla.",
                    )

        missing = _REQUIRED_XLSX_MEMBERS.difference(names)
        if missing:
            _raise_batch_error(
                "INVALID_XLSX_STRUCTURE",
                (
                    "El archivo no contiene la estructura mínima "
                    "de una planilla XLSX."
                ),
            )


def _normalize_header(value: object) -> str:
    if value is None:
        return ""

    return unicodedata.normalize(
        "NFKC",
        str(value),
    ).strip()


def _cell_is_blank(value: object) -> bool:
    return value is None or (
        isinstance(value, str)
        and not value.strip()
    )


def parsear_excel_lotes(
    payload: bytes | bytearray | memoryview,
    *,
    filename: str,
) -> BatchWorkbook:
    """
    Parse a bounded XLSX workbook into the canonical structural schema.

    P2.4A owns structural validation. P2.4B consumes this output for semantic
    row validation without writing anything to PostgreSQL.
    """

    safe_filename = normalizar_nombre_archivo_batch(
        filename,
    )

    if not isinstance(
        payload,
        (bytes, bytearray, memoryview),
    ):
        _raise_batch_error(
            "INVALID_UPLOAD_BODY",
            "El contenido recibido no es un archivo XLSX válido.",
        )

    data = bytes(payload)

    if not data:
        _raise_batch_error(
            "EMPTY_FILE",
            "El archivo Excel está vacío.",
        )

    if len(data) > BATCH_MAX_FILE_BYTES:
        _raise_batch_error(
            "FILE_TOO_LARGE",
            "El archivo Excel excede el tamaño máximo permitido.",
        )

    _preflight_xlsx_container(data)

    workbook = None

    try:
        workbook = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception:
        _raise_batch_error(
            "INVALID_WORKBOOK",
            "No fue posible interpretar la planilla XLSX.",
        )

    try:
        if len(workbook.sheetnames) > BATCH_MAX_SHEETS:
            _raise_batch_error(
                "TOO_MANY_SHEETS",
                "La planilla contiene demasiadas hojas.",
            )

        if BATCH_SHEET_NAME not in workbook.sheetnames:
            _raise_batch_error(
                "MISSING_REQUIRED_SHEET",
                (
                    f"La planilla debe contener la hoja "
                    f"'{BATCH_SHEET_NAME}'."
                ),
            )

        worksheet = workbook[BATCH_SHEET_NAME]

        if (
            worksheet.max_column is None
            or worksheet.max_row is None
        ):
            worksheet.calculate_dimension(
                force=True
            )

        if worksheet.max_column > len(BATCH_COLUMNAS):
            _raise_batch_error(
                "TOO_MANY_COLUMNS",
                "La hoja de importación contiene columnas adicionales.",
            )

        if worksheet.max_row > BATCH_MAX_ROWS + 1:
            _raise_batch_error(
                "TOO_MANY_ROWS",
                (
                    f"La planilla excede el máximo de "
                    f"{BATCH_MAX_ROWS} filas de datos."
                ),
            )

        row_iterator = worksheet.iter_rows(
            min_row=1,
            max_row=BATCH_MAX_ROWS + 2,
            max_col=len(BATCH_COLUMNAS),
        )

        try:
            header_cells = next(row_iterator)
        except StopIteration:
            _raise_batch_error(
                "MISSING_HEADER",
                "La hoja de importación no contiene encabezados.",
            )

        if any(
            cell.data_type == "f"
            for cell in header_cells
        ):
            _raise_batch_error(
                "FORMULA_NOT_ALLOWED",
                "No se admiten fórmulas en la planilla de importación.",
            )

        headers = [
            _normalize_header(cell.value)
            for cell in header_cells
        ]

        if len(set(headers)) != len(headers):
            _raise_batch_error(
                "DUPLICATE_HEADERS",
                "La planilla contiene encabezados duplicados.",
            )

        if headers != BATCH_COLUMNAS:
            _raise_batch_error(
                "INVALID_HEADERS",
                (
                    "Los encabezados de la planilla no coinciden con "
                    "la plantilla oficial de Litoral Trace."
                ),
            )

        rows: list[dict[str, object]] = []
        source_row_numbers: list[int] = []

        for excel_row_number, cells in enumerate(
            row_iterator,
            start=2,
        ):
            if excel_row_number > BATCH_MAX_ROWS + 1:
                if any(
                    not _cell_is_blank(cell.value)
                    for cell in cells
                ):
                    _raise_batch_error(
                        "TOO_MANY_ROWS",
                        (
                            f"La planilla excede el máximo de "
                            f"{BATCH_MAX_ROWS} filas de datos."
                        ),
                    )
                continue

            if any(
                cell.data_type == "f"
                for cell in cells
            ):
                _raise_batch_error(
                    "FORMULA_NOT_ALLOWED",
                    (
                        "No se admiten fórmulas en la planilla "
                        f"de importación (fila {excel_row_number})."
                    ),
                )

            if any(
                cell.data_type == "e"
                for cell in cells
            ):
                _raise_batch_error(
                    "CELL_ERROR",
                    (
                        "La planilla contiene una celda con error "
                        f"de Excel (fila {excel_row_number})."
                    ),
                )

            values = [
                cell.value
                for cell in cells
            ]

            if all(
                _cell_is_blank(value)
                for value in values
            ):
                continue

            rows.append(
                dict(
                    zip(
                        BATCH_COLUMNAS,
                        values,
                        strict=True,
                    )
                )
            )
            source_row_numbers.append(
                excel_row_number
            )

        if not rows:
            _raise_batch_error(
                "NO_DATA_ROWS",
                "La planilla no contiene filas de datos para importar.",
            )

        dataframe = pd.DataFrame(
            rows,
            columns=BATCH_COLUMNAS,
        )

        return BatchWorkbook(
            filename=safe_filename,
            sha256=hashlib.sha256(data).hexdigest(),
            sheet_name=BATCH_SHEET_NAME,
            row_count=len(rows),
            dataframe=dataframe,
            source_row_numbers=tuple(
                source_row_numbers
            ),
        )

    finally:
        if workbook is not None:
            workbook.close()


def _row_error(
    *,
    row: int,
    field: str,
    code: str,
    message: str,
) -> BatchRowError:
    return BatchRowError(
        row=row,
        field=field,
        code=code,
        message=message,
    )


def _normalize_required_text(
    value: object,
    *,
    row: int,
    field: str,
) -> tuple[str | None, list[BatchRowError]]:
    errors: list[BatchRowError] = []

    if value is None:
        errors.append(
            _row_error(
                row=row,
                field=field,
                code="REQUIRED",
                message="El campo es obligatorio.",
            )
        )
        return None, errors

    if not isinstance(value, str):
        errors.append(
            _row_error(
                row=row,
                field=field,
                code="INVALID_TYPE",
                message="El campo debe contener texto.",
            )
        )
        return None, errors

    normalized = unicodedata.normalize(
        "NFKC",
        value,
    ).strip()

    if not normalized:
        errors.append(
            _row_error(
                row=row,
                field=field,
                code="REQUIRED",
                message="El campo es obligatorio.",
            )
        )
        return None, errors

    if _CONTROL_CHARACTER_RE.search(normalized):
        errors.append(
            _row_error(
                row=row,
                field=field,
                code="CONTROL_CHARACTERS",
                message="El campo contiene caracteres de control no permitidos.",
            )
        )
        return None, errors

    normalized = " ".join(
        normalized.split()
    )

    if len(normalized) > BATCH_MAX_TEXT_LENGTH:
        errors.append(
            _row_error(
                row=row,
                field=field,
                code="STRING_TOO_LONG",
                message=(
                    f"El campo excede el máximo de "
                    f"{BATCH_MAX_TEXT_LENGTH} caracteres."
                ),
            )
        )
        return None, errors

    return normalized, errors


def _normalize_required_number(
    value: object,
    *,
    row: int,
    field: str,
) -> tuple[float | None, list[BatchRowError]]:
    errors: list[BatchRowError] = []

    if value is None:
        errors.append(
            _row_error(
                row=row,
                field=field,
                code="REQUIRED",
                message="El campo es obligatorio.",
            )
        )
        return None, errors

    if isinstance(value, bool) or not isinstance(value, Real):
        errors.append(
            _row_error(
                row=row,
                field=field,
                code="INVALID_TYPE",
                message="El campo debe contener un número.",
            )
        )
        return None, errors

    normalized = float(value)

    if not math.isfinite(normalized):
        errors.append(
            _row_error(
                row=row,
                field=field,
                code="NOT_FINITE",
                message="El campo debe contener un número finito.",
            )
        )
        return None, errors

    return normalized, errors


def _validate_semantic_row(
    row_data: dict[str, object],
    *,
    source_row: int,
) -> BatchRowValidation:
    errors: list[BatchRowError] = []

    identificador, field_errors = _normalize_required_text(
        row_data.get("Identificador_Lote"),
        row=source_row,
        field="Identificador_Lote",
    )
    errors.extend(field_errors)

    productor_id, field_errors = _normalize_required_text(
        row_data.get("ID_Proveedor"),
        row=source_row,
        field="ID_Proveedor",
    )
    errors.extend(field_errors)

    producto_forestal, field_errors = _normalize_required_text(
        row_data.get("Producto_Forestal"),
        row=source_row,
        field="Producto_Forestal",
    )
    errors.extend(field_errors)

    hectareas, field_errors = _normalize_required_number(
        row_data.get("Hectareas"),
        row=source_row,
        field="Hectareas",
    )
    errors.extend(field_errors)

    latitud, field_errors = _normalize_required_number(
        row_data.get("Latitud"),
        row=source_row,
        field="Latitud",
    )
    errors.extend(field_errors)

    longitud, field_errors = _normalize_required_number(
        row_data.get("Longitud"),
        row=source_row,
        field="Longitud",
    )
    errors.extend(field_errors)

    volumen_ingresado, field_errors = _normalize_required_number(
        row_data.get("Volumen_Ingresado_Ton"),
        row=source_row,
        field="Volumen_Ingresado_Ton",
    )
    errors.extend(field_errors)

    volumen_exportar, field_errors = _normalize_required_number(
        row_data.get("Volumen_Exportar_Ton"),
        row=source_row,
        field="Volumen_Exportar_Ton",
    )
    errors.extend(field_errors)

    if hectareas is not None and hectareas <= 0.0:
        errors.append(
            _row_error(
                row=source_row,
                field="Hectareas",
                code="MUST_BE_POSITIVE",
                message="La superficie debe ser mayor que cero.",
            )
        )

    if latitud is not None and not (-90.0 <= latitud <= 90.0):
        errors.append(
            _row_error(
                row=source_row,
                field="Latitud",
                code="OUT_OF_RANGE",
                message="La latitud debe estar entre -90 y 90.",
            )
        )

    if longitud is not None and not (-180.0 <= longitud <= 180.0):
        errors.append(
            _row_error(
                row=source_row,
                field="Longitud",
                code="OUT_OF_RANGE",
                message="La longitud debe estar entre -180 y 180.",
            )
        )

    if (
        volumen_ingresado is not None
        and volumen_ingresado < 0.0
    ):
        errors.append(
            _row_error(
                row=source_row,
                field="Volumen_Ingresado_Ton",
                code="MUST_BE_NON_NEGATIVE",
                message="El volumen ingresado no puede ser negativo.",
            )
        )

    if (
        volumen_exportar is not None
        and volumen_exportar < 0.0
    ):
        errors.append(
            _row_error(
                row=source_row,
                field="Volumen_Exportar_Ton",
                code="MUST_BE_NON_NEGATIVE",
                message="El volumen a exportar no puede ser negativo.",
            )
        )

    if (
        volumen_ingresado is not None
        and volumen_exportar is not None
        and volumen_ingresado >= 0.0
        and volumen_exportar >= 0.0
        and volumen_exportar > volumen_ingresado
    ):
        errors.append(
            _row_error(
                row=source_row,
                field="Volumen_Exportar_Ton",
                code="EXPORT_EXCEEDS_INPUT",
                message=(
                    "El volumen a exportar no puede superar "
                    "el volumen ingresado."
                ),
            )
        )

    if errors:
        return BatchRowValidation(
            row=source_row,
            valid=False,
            data=None,
            errors=tuple(errors),
        )

    assert identificador is not None
    assert productor_id is not None
    assert producto_forestal is not None
    assert hectareas is not None
    assert latitud is not None
    assert longitud is not None
    assert volumen_ingresado is not None
    assert volumen_exportar is not None

    return BatchRowValidation(
        row=source_row,
        valid=True,
        data=BatchCanonicalRow(
            source_row=source_row,
            identificador=identificador,
            productor_id=productor_id,
            producto_forestal=producto_forestal,
            hectareas=hectareas,
            latitud=latitud,
            longitud=longitud,
            volumen_ingresado_ton=volumen_ingresado,
            volumen_exportar_ton=volumen_exportar,
        ),
        errors=(),
    )


def validar_dataframe_lotes(
    dataframe: pd.DataFrame,
    *,
    source_row_numbers: tuple[int, ...] | list[int] | None = None,
) -> BatchValidationResult:
    """
    Validate every row without persistence or side effects.

    This is the P2.4B semantic contract consumed by later preview/import APIs.
    """

    if dataframe is None:
        _raise_batch_error(
            "INVALID_DATAFRAME",
            "No se recibió una matriz de datos válida.",
        )

    if list(dataframe.columns) != BATCH_COLUMNAS:
        _raise_batch_error(
            "INVALID_HEADERS",
            (
                "Los encabezados no coinciden con la plantilla "
                "oficial de Litoral Trace."
            ),
        )

    row_count = len(dataframe.index)

    if row_count > BATCH_MAX_ROWS:
        _raise_batch_error(
            "TOO_MANY_ROWS",
            (
                f"La planilla excede el máximo de "
                f"{BATCH_MAX_ROWS} filas de datos."
            ),
        )

    if source_row_numbers is None:
        resolved_source_rows = tuple(
            range(2, row_count + 2)
        )
    else:
        resolved_source_rows = tuple(
            int(value)
            for value in source_row_numbers
        )

    if len(resolved_source_rows) != row_count:
        _raise_batch_error(
            "ROW_MAPPING_MISMATCH",
            "No fue posible mapear las filas a su posición de origen.",
        )

    results: list[BatchRowValidation] = []

    for position, (_, pandas_row) in enumerate(
        dataframe.iterrows()
    ):
        row_payload = {
            column: pandas_row[column]
            for column in BATCH_COLUMNAS
        }

        results.append(
            _validate_semantic_row(
                row_payload,
                source_row=resolved_source_rows[position],
            )
        )

    duplicate_map: dict[str, list[int]] = {}

    for index, row_result in enumerate(results):
        if (
            row_result.data is None
            or not row_result.data.identificador
        ):
            continue

        duplicate_key = row_result.data.identificador.casefold()
        duplicate_map.setdefault(
            duplicate_key,
            [],
        ).append(index)

    duplicate_indexes = {
        index
        for indexes in duplicate_map.values()
        if len(indexes) > 1
        for index in indexes
    }

    if duplicate_indexes:
        rewritten_results: list[BatchRowValidation] = []

        for index, row_result in enumerate(results):
            if index not in duplicate_indexes:
                rewritten_results.append(
                    row_result
                )
                continue

            duplicate_error = _row_error(
                row=row_result.row,
                field="Identificador_Lote",
                code="DUPLICATE_IN_FILE",
                message=(
                    "El identificador del lote está repetido "
                    "dentro de la planilla."
                ),
            )

            rewritten_results.append(
                BatchRowValidation(
                    row=row_result.row,
                    valid=False,
                    data=None,
                    errors=(
                        *row_result.errors,
                        duplicate_error,
                    ),
                )
            )

        results = rewritten_results

    valid_rows = sum(
        1
        for row_result in results
        if row_result.valid
    )
    invalid_rows = len(results) - valid_rows

    return BatchValidationResult(
        valid=invalid_rows == 0,
        total_rows=len(results),
        valid_rows=valid_rows,
        invalid_rows=invalid_rows,
        rows=tuple(results),
    )


def validar_filas_lotes(
    workbook: BatchWorkbook,
) -> BatchValidationResult:
    """Validate the semantic contents of a structurally safe workbook."""

    return validar_dataframe_lotes(
        workbook.dataframe,
        source_row_numbers=workbook.source_row_numbers,
    )


def generar_plantilla_excel() -> bytes:
    """Generate the official XLSX template for batch lot ingestion."""

    df_template = pd.DataFrame(
        columns=BATCH_COLUMNAS,
    )
    df_template.loc[0] = BATCH_FILA_EJEMPLO

    buffer = io.BytesIO()

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl",
    ) as writer:
        df_template.to_excel(
            writer,
            index=False,
            sheet_name=BATCH_SHEET_NAME,
        )

    return buffer.getvalue()


def _polygon_from_point(
    *,
    latitud: float,
    longitud: float,
) -> str:
    """
    Preserve the legacy evidence geometry from already validated coordinates.

    This does not invent coordinates. It deterministically derives the same
    small polygon that the legacy report path used from valid source values.
    """

    delta = 0.01

    return (
        "POLYGON(("
        f"{longitud - delta} {latitud - delta}, "
        f"{longitud + delta} {latitud - delta}, "
        f"{longitud + delta} {latitud + delta}, "
        f"{longitud - delta} {latitud + delta}, "
        f"{longitud - delta} {latitud - delta}"
        "))"
    )


def procesar_lote_masivo(
    df_upload: pd.DataFrame,
) -> tuple[pd.DataFrame, bytes]:
    """
    Generate the legacy evidence ZIP only after strict semantic validation.

    P2.4B removes all permissive numeric/text defaults. Invalid input now fails
    closed with BatchSemanticValidationError and produces no partial artifact.
    """

    resumen_filas: list[dict[str, object]] = []
    zip_buffer = io.BytesIO()

    if df_upload is None or df_upload.empty:
        return (
            pd.DataFrame(resumen_filas),
            zip_buffer.getvalue(),
        )

    validation = validar_dataframe_lotes(
        df_upload,
    )

    if not validation.valid:
        raise BatchSemanticValidationError(
            validation
        )

    with zipfile.ZipFile(
        zip_buffer,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as zip_file:
        for canonical_row in validation.canonical_rows:
            lote_data = {
                "identificador": canonical_row.identificador,
                "productor_id": canonical_row.productor_id,
                "producto_forestal": canonical_row.producto_forestal,
                "hectareas": canonical_row.hectareas,
                "latitud": canonical_row.latitud,
                "longitud": canonical_row.longitud,
                "polygon_wkt": _polygon_from_point(
                    latitud=canonical_row.latitud,
                    longitud=canonical_row.longitud,
                ),
            }

            eval_res = evaluar_compliance_lote(
                lote_data,
                canonical_row.volumen_ingresado_ton,
                canonical_row.volumen_exportar_ton,
            )
            dictamen = eval_res["dictamen"]
            observacion = eval_res["observacion"]
            balance_masas = eval_res["balance_masas"]

            resumen_filas.append(
                {
                    "Lote": canonical_row.identificador,
                    "Proveedor": canonical_row.productor_id,
                    "Producto": canonical_row.producto_forestal,
                    (
                        "Vol. Exportar (Ton)"
                    ): canonical_row.volumen_exportar_ton,
                    "Dictamen": dictamen,
                    "Observación": observacion,
                }
            )

            pdf_bytes = generar_pdf_reporte_bytes(
                lote_data,
                dictamen,
                observacion,
                canonical_row.volumen_ingresado_ton,
                canonical_row.volumen_exportar_ton,
                balance_masas.coeficiente_rendimiento,
            )

            carpeta = (
                f"{dictamen}_"
                f"{canonical_row.productor_id}_"
                f"{canonical_row.identificador}/"
            )

            zip_file.writestr(
                (
                    f"{carpeta}"
                    f"AUDITORIA_{canonical_row.productor_id}.pdf"
                ),
                pdf_bytes,
            )

            if dictamen == "Verde":
                json_data = generar_dds_json_traces_nt(
                    lote_data,
                    canonical_row.volumen_exportar_ton,
                )

                zip_file.writestr(
                    (
                        f"{carpeta}"
                        f"DDS_TRACES_NT_"
                        f"{canonical_row.productor_id}.json"
                    ),
                    json_data.encode("utf-8"),
                )

    return (
        pd.DataFrame(resumen_filas),
        zip_buffer.getvalue(),
    )
