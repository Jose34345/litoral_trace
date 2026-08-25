"""Canonical projection from a confirmed Smart Import mapping.

The canonicalizer is intentionally side-effect free. It transforms only the
selected worksheet columns into LT's existing canonical batch schema. Final
semantic validation and persistence remain owned by the established batch
pipeline.
"""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
import hashlib
import io

import pandas as pd
from openpyxl import load_workbook

from litoral_trace.services.batch import (
    BATCH_COLUMNAS,
    BATCH_MAX_ROWS,
    BatchWorkbook,
    _preflight_xlsx_container,
    normalizar_nombre_archivo_batch,
)

from .contracts import DatasetCandidate
from .engine import SMART_MAX_FILE_BYTES, SmartImportError


@dataclass(frozen=True)
class ConfirmedMapping:
    """One user-confirmed source-to-canonical mapping."""

    source_index: int
    source_column: str
    canonical_field: str


class SmartCanonicalizationError(SmartImportError):
    """Raised when confirmed mappings cannot produce a valid canonical table."""


def _raise(code: str, detail: str) -> None:
    raise SmartCanonicalizationError(code=code, detail=detail)


def default_confirmed_mapping(candidate: DatasetCandidate) -> tuple[ConfirmedMapping, ...]:
    """Build the strongest proposed projection for each canonical target.

    This helper is intended to pre-populate a future confirmation UI. Callers
    are still responsible for obtaining human confirmation for mappings whose
    discovery status requires it before importing business data.
    """

    best_by_target = {}
    for mapping in candidate.mappings:
        target = mapping.decision.canonical_field
        if target is None or target not in BATCH_COLUMNAS:
            continue
        current = best_by_target.get(target)
        if current is None or mapping.decision.confidence > current.decision.confidence:
            best_by_target[target] = mapping

    return tuple(
        ConfirmedMapping(
            source_index=mapping.source_index,
            source_column=mapping.source_column,
            canonical_field=target,
        )
        for target in BATCH_COLUMNAS
        if (mapping := best_by_target.get(target)) is not None
    )


def canonicalize_workbook(
    payload: bytes | bytearray | memoryview,
    *,
    filename: str,
    candidate: DatasetCandidate,
    mappings: Sequence[ConfirmedMapping],
    max_rows: int = BATCH_MAX_ROWS,
) -> BatchWorkbook:
    """Project one selected dataset into the existing LT canonical schema.

    Extra source columns are ignored by design. All eight canonical fields must
    be explicitly mapped; no values are invented. The returned `BatchWorkbook`
    is intentionally bounded by the current batch validator's row contract so
    it can be passed directly to `validar_filas_lotes` without weakening any
    existing safety invariant. Large-job/chunk ingestion is a later gate.
    """

    safe_filename = normalizar_nombre_archivo_batch(filename)
    if not isinstance(payload, (bytes, bytearray, memoryview)):
        _raise("INVALID_UPLOAD_BODY", "El contenido recibido no es un XLSX válido.")

    data = bytes(payload)
    if not data:
        _raise("EMPTY_FILE", "El archivo Excel está vacío.")
    if len(data) > SMART_MAX_FILE_BYTES:
        _raise("SMART_FILE_TOO_LARGE", "El archivo excede el límite seguro de Smart Import.")
    if max_rows <= 0 or max_rows > BATCH_MAX_ROWS:
        _raise(
            "INVALID_ROW_LIMIT",
            f"Smart Import V1 admite hasta {BATCH_MAX_ROWS} filas por importación canónica.",
        )

    targets = [mapping.canonical_field for mapping in mappings]
    duplicates = sorted({target for target in targets if targets.count(target) > 1})
    if duplicates:
        _raise(
            "DUPLICATE_CANONICAL_MAPPING",
            "Hay más de una columna fuente asignada al mismo campo canónico.",
        )

    unknown = sorted(set(targets) - set(BATCH_COLUMNAS))
    if unknown:
        _raise(
            "UNKNOWN_CANONICAL_FIELD",
            "El mapping contiene campos que no pertenecen al esquema canónico de lotes.",
        )

    missing = [column for column in BATCH_COLUMNAS if column not in targets]
    if missing:
        _raise(
            "MISSING_REQUIRED_MAPPING",
            "Faltan campos obligatorios para construir el dataset canónico: "
            + ", ".join(missing),
        )

    _preflight_xlsx_container(data)
    workbook = None
    try:
        workbook = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        if candidate.sheet_name not in workbook.sheetnames:
            _raise(
                "MISSING_SELECTED_SHEET",
                "La hoja seleccionada ya no existe en el workbook.",
            )

        worksheet = workbook[candidate.sheet_name]
        mapping_by_index = {mapping.source_index: mapping for mapping in mappings}
        source_indices = sorted(mapping_by_index)
        if not source_indices:
            _raise("EMPTY_MAPPING", "No hay columnas confirmadas para importar.")

        max_source_index = max(source_indices)
        if max_source_index >= (worksheet.max_column or 0):
            _raise(
                "SOURCE_COLUMN_OUT_OF_RANGE",
                "El mapping hace referencia a una columna que no existe en la hoja.",
            )

        rows: list[dict[str, object]] = []
        source_row_numbers: list[int] = []

        for row_number, cells in enumerate(
            worksheet.iter_rows(
                min_row=candidate.first_data_row,
                max_col=max_source_index + 1,
            ),
            start=candidate.first_data_row,
        ):
            projected: dict[str, object] = {}
            non_blank = False
            for source_index in source_indices:
                cell = cells[source_index]
                value = cell.value
                if value is not None and not (isinstance(value, str) and not value.strip()):
                    non_blank = True
                if cell.data_type == "f":
                    _raise(
                        "FORMULA_IN_MAPPED_COLUMN",
                        f"No se admiten fórmulas en columnas mapeadas (fila {row_number}).",
                    )
                if cell.data_type == "e":
                    _raise(
                        "CELL_ERROR_IN_MAPPED_COLUMN",
                        f"La planilla contiene una celda con error en una columna mapeada (fila {row_number}).",
                    )
                mapping = mapping_by_index[source_index]
                projected[mapping.canonical_field] = value

            if not non_blank:
                continue

            rows.append({column: projected.get(column) for column in BATCH_COLUMNAS})
            source_row_numbers.append(row_number)
            if len(rows) > max_rows:
                _raise(
                    "SMART_TOO_MANY_DATA_ROWS",
                    f"El dataset supera el máximo temporal de {max_rows} filas por importación Smart V1.",
                )

        if not rows:
            _raise("NO_DATA_ROWS", "La tabla seleccionada no contiene filas para importar.")

        dataframe = pd.DataFrame(rows, columns=BATCH_COLUMNAS)
        return BatchWorkbook(
            filename=safe_filename,
            sha256=hashlib.sha256(data).hexdigest(),
            sheet_name=candidate.sheet_name,
            row_count=len(rows),
            dataframe=dataframe,
            source_row_numbers=tuple(source_row_numbers),
        )
    except SmartImportError:
        raise
    except Exception as exc:
        raise SmartCanonicalizationError(
            code="SMART_CANONICALIZATION_FAILED",
            detail="No fue posible proyectar el dataset al esquema canónico de Litoral Trace.",
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()
