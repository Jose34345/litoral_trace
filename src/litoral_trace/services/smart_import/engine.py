"""Workbook discovery engine for Smart Excel Import V1.

This module deliberately does not persist data. It discovers likely tabular
regions and proposes deterministic mappings to Litoral Trace's canonical lot
schema. The existing semantic validator and atomic persistence service remain
the final authority before database writes.
"""

from __future__ import annotations

from dataclasses import replace
import hashlib
import io
from typing import Any, BinaryIO

from openpyxl import load_workbook

from litoral_trace.services.batch import (
    _preflight_xlsx_container,
    normalizar_nombre_archivo_batch,
)

from .aliases import LOTES_CANONICAL_FIELDS
from .contracts import (
    ColumnMapping,
    DatasetCandidate,
    MappingStatus,
    SmartWorkbookAnalysis,
)
from .matcher import map_source_column, resolve_duplicate_targets


SMART_MAX_FILE_BYTES = 25 * 1024 * 1024
SMART_MAX_SHEETS = 20
SMART_HEADER_SCAN_ROWS = 25
SMART_SAMPLE_ROWS = 24
SMART_MAX_DISCOVERY_COLUMNS = 256
SMART_MIN_HEADER_CELLS = 2
SMART_MIN_DATASET_SCORE = 0.18


class SmartImportError(ValueError):
    """Safe user-originated Smart Import failure."""

    def __init__(self, code: str, detail: str) -> None:
        self.code = code
        self.detail = detail
        super().__init__(detail)


def _error(code: str, detail: str) -> None:
    raise SmartImportError(code=code, detail=detail)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _row_values(worksheet: Any, row_number: int, max_column: int) -> list[Any]:
    return [
        cell.value
        for cell in next(
            worksheet.iter_rows(
                min_row=row_number,
                max_row=row_number,
                min_col=1,
                max_col=max_column,
            )
        )
    ]


def _column_samples(
    worksheet: Any,
    *,
    header_row: int,
    source_index: int,
) -> tuple[Any, ...]:
    values: list[Any] = []
    start = header_row + 1
    end = min(
        worksheet.max_row or start,
        header_row + SMART_SAMPLE_ROWS,
    )

    if end < start:
        return ()

    for row in worksheet.iter_rows(
        min_row=start,
        max_row=end,
        min_col=source_index + 1,
        max_col=source_index + 1,
    ):
        value = row[0].value
        if not _is_blank(value):
            values.append(value)
        if len(values) >= 8:
            break

    return tuple(values)


def _header_quality(mappings: tuple[ColumnMapping, ...]) -> tuple[float, int, int]:
    usable = [
        mapping
        for mapping in mappings
        if mapping.decision.canonical_field is not None
    ]
    auto_or_confirm = [
        mapping
        for mapping in usable
        if mapping.decision.status in {MappingStatus.AUTO, MappingStatus.CONFIRM}
    ]

    mapped_targets = {
        mapping.decision.canonical_field
        for mapping in usable
        if mapping.decision.canonical_field is not None
    }
    required_targets = {field.name for field in LOTES_CANONICAL_FIELDS if field.required}
    coverage = len(mapped_targets & required_targets) / max(1, len(required_targets))
    confidence = (
        sum(mapping.decision.confidence for mapping in usable) / len(usable)
        if usable
        else 0.0
    )
    strong_ratio = len(auto_or_confirm) / max(1, len(usable))

    score = (0.58 * coverage) + (0.27 * confidence) + (0.15 * strong_ratio)
    return score, len(mapped_targets), len(auto_or_confirm)


def _candidate_for_header(
    worksheet: Any,
    *,
    sheet_name: str,
    header_row: int,
    max_column: int,
) -> DatasetCandidate | None:
    headers = _row_values(worksheet, header_row, max_column)
    populated = [value for value in headers if not _is_blank(value)]
    if len(populated) < SMART_MIN_HEADER_CELLS:
        return None

    mappings = tuple(
        map_source_column(
            header,
            _column_samples(
                worksheet,
                header_row=header_row,
                source_index=index,
            ),
            source_index=index,
        )
        for index, header in enumerate(headers)
        if not _is_blank(header)
    )
    mappings = resolve_duplicate_targets(mappings)

    score, mapped_count, strong_count = _header_quality(mappings)
    if mapped_count < 2 or score < SMART_MIN_DATASET_SCORE:
        return None

    mapped_fields = {
        mapping.decision.canonical_field
        for mapping in mappings
        if mapping.decision.canonical_field is not None
    }
    missing_required = tuple(
        field.name
        for field in LOTES_CANONICAL_FIELDS
        if field.required and field.name not in mapped_fields
    )

    # Slightly favor candidates with multiple strong deterministic mappings.
    score = min(1.0, score + min(0.06, strong_count * 0.008))

    return DatasetCandidate(
        sheet_name=sheet_name,
        header_row=header_row,
        first_data_row=header_row + 1,
        estimated_rows=max(0, (worksheet.max_row or header_row) - header_row),
        estimated_columns=max_column,
        score=score,
        mappings=mappings,
        missing_required_fields=missing_required,
    )


class SmartImportEngine:
    """Discover likely LT datasets in a heterogeneous XLSX workbook."""

    def analyze(
        self,
        payload: bytes | bytearray | memoryview,
        *,
        filename: str,
    ) -> SmartWorkbookAnalysis:
        safe_filename = normalizar_nombre_archivo_batch(filename)

        if not isinstance(payload, (bytes, bytearray, memoryview)):
            _error("INVALID_UPLOAD_BODY", "El contenido recibido no es un XLSX válido.")

        data = bytes(payload)
        if not data:
            _error("EMPTY_FILE", "El archivo Excel está vacío.")
        if len(data) > SMART_MAX_FILE_BYTES:
            _error(
                "SMART_FILE_TOO_LARGE",
                "El archivo excede el límite seguro de Smart Import (25 MB).",
            )

        # Reuse LT's hardened ZIP/XML preflight before openpyxl sees the bytes.
        _preflight_xlsx_container(data)

        workbook = None
        try:
            workbook = load_workbook(
                io.BytesIO(data),
                read_only=True,
                data_only=False,
                keep_links=False,
            )

            if len(workbook.sheetnames) > SMART_MAX_SHEETS:
                _error(
                    "SMART_TOO_MANY_SHEETS",
                    f"El workbook excede el máximo de {SMART_MAX_SHEETS} hojas analizables.",
                )

            candidates: list[DatasetCandidate] = []
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                max_row = worksheet.max_row or 0
                max_column = worksheet.max_column or 0
                if max_row <= 0 or max_column <= 0:
                    continue

                discovery_columns = min(max_column, SMART_MAX_DISCOVERY_COLUMNS)
                scan_rows = min(max_row, SMART_HEADER_SCAN_ROWS)
                sheet_candidates = [
                    candidate
                    for header_row in range(1, scan_rows + 1)
                    if (
                        candidate := _candidate_for_header(
                            worksheet,
                            sheet_name=sheet_name,
                            header_row=header_row,
                            max_column=discovery_columns,
                        )
                    )
                    is not None
                ]

                # Keep only the best plausible tabular region per sheet in V1.
                if sheet_candidates:
                    candidates.append(
                        max(sheet_candidates, key=lambda candidate: candidate.score)
                    )

            candidates.sort(key=lambda candidate: candidate.score, reverse=True)
            return SmartWorkbookAnalysis(
                filename=safe_filename,
                sha256=hashlib.sha256(data).hexdigest(),
                sheet_names=tuple(workbook.sheetnames),
                candidates=tuple(candidates),
            )
        except SmartImportError:
            raise
        except Exception as exc:
            raise SmartImportError(
                code="SMART_INVALID_WORKBOOK",
                detail="No fue posible analizar la estructura del workbook.",
            ) from exc
        finally:
            if workbook is not None:
                workbook.close()
